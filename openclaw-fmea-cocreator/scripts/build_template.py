"""One-shot script to (re)build template.xlsx for M2.

Usage:
    python3 openclaw-fmea-cocreator/scripts/build_template.py
"""
from __future__ import annotations

import shutil
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

SKILL_DIR = Path(__file__).resolve().parent.parent
TEMPLATE_PATH = SKILL_DIR / "template.xlsx"

FMEA_HEADERS = [
    "序号", "Scope path", "Leaf 节点", "Analysis object", "Function or requirement",
    "P-Diagram 锚点", "Failure mode", "Failure mode canonical", "Failure effect",
    "S", "Cause or mechanism", "O", "Current controls (prevention)",
    "Current controls (detection)", "D", "RPN", "Recommended actions",
    "Owner", "Target date",
    "改进后 S", "改进后 O", "改进后 D", "改进后 RPN",
    "Evidence grade", "Confidence", "Confidence breakdown",
    "Multi-role corroborated", "Rating history",
    "Needs human confirmation", "Source traces", "AI 打分推导依据",
]
assert len(FMEA_HEADERS) == 31, "Header count must equal 31"

HEADER_FILL = PatternFill("solid", fgColor="333333")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SERIAL_FILL = PatternFill("solid", fgColor="D9E1F2")
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
TEXT_ALIGN = Alignment(horizontal="left", vertical="top", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
NUMERIC_COLUMNS = {11, 13, 16, 17, 21, 22, 23, 24, 26}


def build():
    wb = Workbook()

    cover = wb.active
    cover.title = "封面"
    cover["B2"] = "<模块名> <FMEA类型>分析报告"
    cover["B6"] = "模块"
    cover["B7"] = "关键功能/指标"
    cover["B8"] = "采用的方法学"
    cover["B9"] = "范围/Scopes"
    cover["B10"] = "数据来源"
    cover["B11"] = "生成日期"
    cover["B12"] = "版本"
    cover["B14"] = "覆盖摘要"
    cover["B15"] = "Hierarchy 节点数"
    cover["B16"] = "Coverage gaps 行数"
    cover["B17"] = "证据等级分布"
    cover["B18"] = "置信度分布"
    cover["B19"] = "评审导引: 先看 Sheet 4 待确认队列"

    main = wb.create_sheet("FMEA主表")
    for col_idx, header in enumerate(FMEA_HEADERS, start=2):
        cell = main.cell(row=2, column=col_idx)
        cell.value = header
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGN
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
    main.row_dimensions[2].height = 32

    main.cell(row=3, column=17).value = "=J3*L3*O3"
    main.cell(row=3, column=24).value = "=U3*V3*W3"
    for col_idx in range(2, 33):
        cell = main.cell(row=3, column=col_idx)
        cell.border = THIN_BORDER
        if col_idx == 2:
            cell.fill = SERIAL_FILL
            cell.font = Font(bold=True)
            cell.alignment = CENTER_ALIGN
        elif col_idx in NUMERIC_COLUMNS:
            cell.alignment = CENTER_ALIGN
        else:
            cell.alignment = TEXT_ALIGN

    width_map = {
        2: 6, 3: 20, 4: 16, 5: 18, 6: 22, 7: 24, 8: 22, 9: 22, 10: 30, 11: 5,
        12: 30, 13: 5, 14: 22, 15: 22, 16: 5, 17: 8, 18: 28, 19: 12, 20: 14,
        21: 7, 22: 7, 23: 7, 24: 8, 25: 16, 26: 10, 27: 26, 28: 14, 29: 24,
        30: 12, 31: 26, 32: 26,
    }
    for col_idx, w in width_map.items():
        main.column_dimensions[get_column_letter(col_idx)].width = w

    rules = wb.create_sheet("评分准则参考")
    rules["B2"] = "Severity (S)"
    rules["B3"] = "1=无影响 ... 10=安全/法规红线"
    rules["B5"] = "Occurrence (O)"
    rules["B6"] = "1=极低 ... 10=极频繁"
    rules["B8"] = "Detection (D)"
    rules["B9"] = "1=必然检出 ... 10=完全无法检出"
    rules["B11"] = "评分准则随企业能力而变,本表仅为概念参考。"

    gaps = wb.create_sheet("覆盖盲区与待确认队列")
    gaps["B2"] = "覆盖盲区 (coverage_gaps)"
    for col_idx, h in enumerate(["leaf_id", "role", "axis_combo", "severity_estimate"], start=2):
        c = gaps.cell(row=3, column=col_idx, value=h)
        c.font = Font(bold=True)
    gaps["B10"] = "待确认队列 (confirmation_queue)"
    for col_idx, h in enumerate(["row_id", "leaf_id", "failure_mode", "evidence_grade", "confidence", "rpn", "confidence × rpn"], start=2):
        c = gaps.cell(row=11, column=col_idx, value=h)
        c.font = Font(bold=True)

    struct = wb.create_sheet("结构与P-Diagram")
    struct["B2"] = "Hierarchy"
    struct["B3"] = "(由 build_workbook.py 在生成时填入树状缩进)"
    struct["B10"] = "P-Diagrams"
    struct["B11"] = "(每个子系统一段:scope_id / 6 轴明细)"

    TEMPLATE_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(TEMPLATE_PATH)


if __name__ == "__main__":
    build()
    print(f"wrote {TEMPLATE_PATH}")
