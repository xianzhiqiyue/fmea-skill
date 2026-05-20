from __future__ import annotations

import argparse
import json
import re
from collections import Counter, defaultdict
from copy import copy
from dataclasses import asdict, dataclass, field
from datetime import date
from pathlib import Path
from typing import Any

from retrieve_cases import (
    ALIASES,
    DATA_ROOT,
    RELATED_MODULES,
    Match,
    collect_matches,
    row_text,
    tokenize,
)


SKILL_DIR = Path(__file__).resolve().parent.parent
PROJECT_ROOT = SKILL_DIR.parent
PACKAGED_TEMPLATE_PATH = SKILL_DIR / "template.xlsx"
REPO_TEMPLATE_PATH = PROJECT_ROOT / "template.xlsx"
DEFAULT_TEMPLATE_PATH = PACKAGED_TEMPLATE_PATH if PACKAGED_TEMPLATE_PATH.exists() else REPO_TEMPLATE_PATH
VALID_FMEA_TYPES = {"AFMEA", "SFMEA", "DFMEA", "PFMEA"}
TEMPLATE_PATHS_BY_TYPE = {
    "AFMEA": SKILL_DIR / "afmea_template.xlsx",
    "SFMEA": SKILL_DIR / "sfmea_template.xlsx",
    "DFMEA": PACKAGED_TEMPLATE_PATH,
    "PFMEA": SKILL_DIR / "pfmea_template.xlsx",
}
FMEA_TYPE_METADATA = {
    "AFMEA": {
        "subtitle": "Application FMEA for {module_en} - Product Lifecycle Approach",
        "indicator_fallback": "{module}应用生命周期、环境、运输、安装、操作和维护场景",
        "standard_note": "AIAG-VDA FMEA Handbook（第1版）七步法；生命周期/场景/控制点展开",
        "default_min_rows": 28,
    },
    "SFMEA": {
        "subtitle": "System FMEA for {module_en} - System Boundary and Interface Approach",
        "indicator_fallback": "{module}系统边界、子系统接口、功能链、能量/物料/信息流",
        "standard_note": "AIAG-VDA FMEA Handbook（第1版）七步法；系统边界/接口/功能链展开",
        "default_min_rows": 25,
    },
    "DFMEA": {
        "subtitle": "Design FMEA for {module_en} - Part-level System Approach",
        "indicator_fallback": "{module}关键功能、接口、零件、材料、器件与设计约束",
        "standard_note": "AIAG-VDA FMEA Handbook（第1版）七步法；系统级边界/接口/零件级展开强约束",
        "default_min_rows": 36,
    },
    "PFMEA": {
        "subtitle": "Process FMEA for {module_en} - Process Step and Control Plan Approach",
        "indicator_fallback": "{module}工艺流程、过程参数、设备工装、检验控制与后工序影响",
        "standard_note": "AIAG-VDA FMEA Handbook（第1版）七步法；过程步骤/工艺参数/控制计划展开",
        "default_min_rows": 30,
    },
}
ALLOWED_THEMES = {"dfmea_sample_data", "knowledge_base_template"}
DEFAULT_SCOPE_MIN_ROWS = 4
COVERAGE_GAP_GUIDEWORDS: list[dict[str, str]] = [
    {
        "category": "功能丧失",
        "failure_mode": "核心功能完全丧失",
        "effect": "相关功能不可用,导致任务中断、验收失败或后工序停滞",
        "cause": "关键输入、执行单元、供电/气路/通信或控制链路中断",
        "controls": "待确认是否已有上电自检、功能测试、报警联锁或出厂验证",
        "action": "补充关键功能自检、失效注入验证、异常停机策略和责任人",
        "severity": "8",
        "occurrence": "4",
        "detection": "5",
    },
    {
        "category": "功能退化",
        "failure_mode": "功能性能退化或输出不足",
        "effect": "性能指标偏离要求,造成实验结果失真、效率下降或客户重复调试",
        "cause": "裕量不足、参数漂移、磨损、校准偏移或过程窗口变窄",
        "controls": "待确认是否已有性能边界测试、趋势监测、校准记录或 SPC 控制",
        "action": "补充性能边界验证、趋势预警、校准周期和关键参数上下限",
        "severity": "7",
        "occurrence": "5",
        "detection": "5",
    },
    {
        "category": "间歇性功能",
        "failure_mode": "间歇性失效或偶发不稳定",
        "effect": "现场难以复现,导致误判、重复返修、任务中断或服务成本上升",
        "cause": "接触不良、热漂移、振动松动、边界状态时序竞争或日志不足",
        "controls": "待确认是否已有长稳测试、振动/温升覆盖、日志记录和异常复盘",
        "action": "增加长时间运行、边界循环、日志抓取、连接防松和复现判据",
        "severity": "7",
        "occurrence": "5",
        "detection": "6",
    },
    {
        "category": "非预期功能",
        "failure_mode": "非预期动作或错误状态切换",
        "effect": "设备在错误时机动作,造成样品/工件损伤、安全风险或流程混乱",
        "cause": "状态机覆盖不足、互锁缺失、误触发、配置错用或异常恢复逻辑不完整",
        "controls": "待确认是否已有独立互锁、权限限制、状态覆盖测试和异常恢复测试",
        "action": "补充状态覆盖矩阵、独立联锁、权限控制、异常恢复和失效注入测试",
        "severity": "8",
        "occurrence": "4",
        "detection": "6",
    },
    {
        "category": "错误输出/误判",
        "failure_mode": "输出错误、判断错误或数据张冠李戴",
        "effect": "后工序基于错误结果继续流转,造成误放行、实验结论错误或追溯失效",
        "cause": "传感/测量误差、阈值设置不当、数据绑定错误、缓存未刷新或算法误判",
        "controls": "待确认是否已有双通道校验、扫码绑定、边界样本验证和异常数据拦截",
        "action": "增加数据绑定校验、阈值锁定、异常值拦截、复核逻辑和追溯记录",
        "severity": "8",
        "occurrence": "4",
        "detection": "7",
    },
    {
        "category": "接口失配",
        "failure_mode": "接口匹配不良或边界责任不清",
        "effect": "集成、安装或联调阶段出现连接失败、通信异常、泄漏、松脱或责任争议",
        "cause": "接口控制文件缺失、公差链/协议/接地/密封/流量边界定义不足",
        "controls": "待确认是否已有接口矩阵、边界样机验证、联调清单和变更闭环",
        "action": "补齐接口矩阵、边界条件、关键接口测试、变更评审和交接验收",
        "severity": "7",
        "occurrence": "5",
        "detection": "5",
    },
    {
        "category": "环境应力",
        "failure_mode": "环境应力下失效",
        "effect": "温湿度、EMC、电源、振动或污染导致性能漂移、误报警或停机",
        "cause": "环境裕量不足、客户现场边界未定义、降额不足或环境监测缺失",
        "controls": "待确认是否已有环境适应性验证、降额清单、环境监测和现场预检",
        "action": "增加环境边界验证、降额复核、环境超限联锁和客户现场条件确认",
        "severity": "7",
        "occurrence": "4",
        "detection": "6",
    },
    {
        "category": "老化/磨损",
        "failure_mode": "寿命件老化、磨损或疲劳失效",
        "effect": "全寿命后期可靠性下降,导致重复停机、维护成本上升或关键功能失效",
        "cause": "寿命模型不足、材料老化、疲劳、插拔/运动次数超限或维护周期不清",
        "controls": "待确认是否已有寿命试验、寿命计数、维护提醒和备件策略",
        "action": "建立寿命件清单、寿命计数、维护提醒、加速寿命验证和备件策略",
        "severity": "7",
        "occurrence": "5",
        "detection": "6",
    },
    {
        "category": "误操作/维护错误",
        "failure_mode": "误操作、误维护或参数误设",
        "effect": "客户现场或维修后出现功能异常、安全保护误动作、数据失真或返修",
        "cause": "SOP 不清、权限/提示不足、防错缺失、维护复位步骤遗漏或培训不可验证",
        "controls": "待确认是否已有权限控制、操作确认、维护清单、培训记录和日志追踪",
        "action": "增加防错设计、权限分级、关键步骤二次确认、维护闭环和日志追溯",
        "severity": "7",
        "occurrence": "5",
        "detection": "5",
    },
    {
        "category": "检验逃逸",
        "failure_mode": "检验/测试覆盖不足导致缺陷逃逸",
        "effect": "缺陷在后工序或客户现场暴露,造成误放行、返工、投诉或安全风险",
        "cause": "测试覆盖不足、抽样方案不足、治具/量具状态异常、阈值未锁定或 MSA 不足",
        "controls": "待确认是否已有测试覆盖矩阵、治具点检、MSA、全检/抽检规则和放行审批",
        "action": "补充测试覆盖矩阵、治具自检、MSA 复核、判定阈值锁定和逃逸复盘",
        "severity": "8",
        "occurrence": "4",
        "detection": "8",
    },
]

FIELD_ALIASES = {
    "analysis_object": ["模块/零件", "零件名称", "子系统/功能模块", "子系统/组件", "子系统/部件", "模块", "关联项目/产品"],
    "function": [
        "功能及要求",
        "功能要求",
        "功能描述",
        "标准化功能项",
        "产品寿命周期应用任务和环境剖面",
        "策划和控制内容",
    ],
    "parameter_indicators": ["参数指标性能", "参数指标", "性能指标"],
    "failure_mode": ["潜在失效模式", "失效模式 (AI分类)"],
    "effect": [
        "失效影响（后果）",
        "潜在失效后果 (客户/后工序)",
        "潜在失效后果（客户/后工序）",
        "失效的潜在后果（对客户或后工序的影响）",
        "失效后果 (S)",
    ],
    "severity": ["严重度\nS", "严重度 S", "严重度 (S)", "S"],
    "cause": ["潜在失效起因/机理", "潜在失效原因（机理）", "潜在失效原因", "根本原因分析 (Cause)"],
    "occurrence": ["频度\nO", "频度 O", "发生频次 (O)", "O"],
    "current_controls": [
        "现行设计控制 (预防/探测)",
        "现行控制措施",
        "现行控制方法",
        "现行设计/过程控制措施",
        "现行控制措施",
        "现行预防措施",
        "现行探测控制",
    ],
    "detection": ["探测度\nD", "探测度 D", "可探测度 (D)", "D"],
    "rpn": ["RPN", "初始 RPN"],
    "recommended_actions": [
        "建议改进措施 (控制/预防)",
        "建议的控制措施 (改进方案)",
        "建议措施",
        "建议的预防/探测措施",
    ],
    "rating_basis": ["AI打分推导依据", "评分依据", "打分依据"],
    "post_action_severity": ["改进后S", "措施后 S", "改进后 S", "新S"],
    "post_action_occurrence": ["改进后O", "措施后 O", "改进后 O", "新O"],
    "post_action_detection": ["改进后D", "措施后 D", "改进后 D", "新D"],
    "post_action_rpn": ["改进后RPN", "措施后 RPN", "改进后 RPN", "新RPN"],
    "owner": ["措施负责人", "责任人", "负责人"],
    "target_date": ["完成时间", "计划完成时间", "结束时间"],
}

STOPWORDS = {
    "客户",
    "后工序",
    "系统",
    "功能",
    "要求",
    "建议",
    "设计",
    "控制",
    "改进",
    "方案",
    "模块",
    "项目",
    "产品",
    "组件",
    "部分",
    "过程",
    "进行",
    "用于",
    "增加",
    "导致",
    "无法",
    "工作",
    "模式",
    "分析",
    "输出",
    "表格",
}

SCOPE_PROFILES: dict[str, list[dict[str, Any]]] = {
    "变温系统": [
        {
            "name": "压缩机制冷子系统",
            "keywords": ["压缩机", "冷媒", "气液分离器", "蒸发器", "毛细管", "高压", "继电器", "422", "VX3244", "1028", "启动", "时序"],
            "min_hits": 2,
            "max_terms": 10,
        },
        {
            "name": "液氮蒸发子系统",
            "keywords": ["液氮罐", "主体组件", "排气管组件", "波纹管", "真空", "PTFE", "特氟龙", "换热盘管", "加热棒", "PT100", "航插", "麦拉膜", "G10"],
            "min_hits": 2,
            "max_terms": 10,
        },
    ],
    "自动进样器": [
        {
            "name": "运动与抓取子系统",
            "keywords": ["气缸", "滑块", "夹爪", "储样筒", "进样筒", "转盘", "样品管", "扭簧", "导轨", "定位销"],
            "min_hits": 2,
            "max_terms": 10,
        },
        {
            "name": "检测与气路控制子系统",
            "keywords": ["检测", "光电", "传感器", "对射", "气路", "电磁阀", "减压阀", "缓冲气", "压力开关", "气压", "供电", "接口"],
            "min_hits": 2,
            "max_terms": 10,
        },
    ],
    "调谐单元": [
        {
            "name": "射频调谐与匹配子系统",
            "keywords": ["射频", "调谐", "匹配", "电容", "打火", "线圈", "焊接", "磁化率", "梯度线圈"],
            "min_hits": 2,
            "max_terms": 10,
        },
        {
            "name": "机械传动与限位子系统",
            "keywords": ["电机", "编码器", "限位", "离合器", "连杆", "齿轮", "减速机", "扭矩"],
            "min_hits": 2,
            "max_terms": 10,
        },
        {
            "name": "EMC与算法控制子系统",
            "keywords": ["EMC", "CAN", "RS422", "MCU", "光耦", "搜索", "陷波", "扫掠", "算法"],
            "min_hits": 2,
            "max_terms": 10,
        },
    ],
    "收发机": [
        {
            "name": "发射链路子系统",
            "keywords": ["发射", "上变频", "调幅", "功放驱动", "衰减", "功率", "泄漏"],
            "min_hits": 2,
            "max_terms": 10,
        },
        {
            "name": "接收与采集子系统",
            "keywords": ["接收", "LNA", "ADC", "时钟", "噪声", "解调", "增益"],
            "min_hits": 2,
            "max_terms": 10,
        },
        {
            "name": "频率合成与连接子系统",
            "keywords": ["混频", "本振", "频率合成", "背板", "连接器", "同轴", "杂散"],
            "min_hits": 2,
            "max_terms": 10,
        },
    ],
    "前置放大器": [
        {
            "name": "T/R与保护子系统",
            "keywords": ["T/R", "PIN", "高功率", "保护", "限幅", "功率检测", "联动跳闸", "VSWR", "驻波", "复位", "热管理"],
            "min_hits": 2,
            "max_terms": 12,
        },
        {
            "name": "低噪声放大与接口子系统",
            "keywords": ["LNA", "增益", "振荡", "N-K", "BNC", "接口", "回波损耗"],
            "min_hits": 2,
            "max_terms": 10,
        },
        {
            "name": "控制逻辑与供电子系统",
            "keywords": ["UNWORK", "OBSERVE", "Watchdog", "DB-37", "供电", "串扰", "模式切换"],
            "min_hits": 2,
            "max_terms": 10,
        },
    ],
    "射频功放": [
        {
            "name": "异常保护与联锁子系统",
            "keywords": ["保护", "联锁", "比较器", "切断", "驻波", "反射功率", "异常检测"],
            "min_hits": 2,
            "max_terms": 10,
        },
        {
            "name": "放大与通道切换子系统",
            "keywords": ["放大", "增益", "双通道", "切换", "热切换", "继电器", "功率管"],
            "min_hits": 2,
            "max_terms": 10,
        },
        {
            "name": "热管理与复位子系统",
            "keywords": ["热管理", "温度", "风扇", "热斑", "复位", "Auto-Recovery", "散热"],
            "min_hits": 2,
            "max_terms": 10,
        },
    ],
    "进样筒": [
        {
            "name": "结构与磁兼容子系统",
            "keywords": ["磁性", "无磁", "结构件", "清洗", "磁导率", "铁屑"],
            "min_hits": 2,
            "max_terms": 10,
        },
        {
            "name": "气动与升降子系统",
            "keywords": ["气动", "气源", "缓冲气罐", "软着陆", "升降", "卡阻", "内壁"],
            "min_hits": 2,
            "max_terms": 10,
        },
        {
            "name": "旋转与控制子系统",
            "keywords": ["转子", "旋转", "50Hz", "PID", "光电传感器", "气流阀门"],
            "min_hits": 2,
            "max_terms": 10,
        },
    ],
    "匀场单元": [
        {
            "name": "匀场线圈结构子系统",
            "keywords": ["匀场线圈", "几何形状", "固化", "环氧", "真空灌封", "节点松动"],
            "min_hits": 2,
            "max_terms": 10,
        },
        {
            "name": "恒流驱动与连接子系统",
            "keywords": ["SDB", "恒流", "DAC", "基准电压", "DB50", "连接", "检流电阻"],
            "min_hits": 2,
            "max_terms": 10,
        },
        {
            "name": "软件与耦合管理子系统",
            "keywords": ["算法", "通信", "EEPROM", "Watchdog", "互感", "低温匀场", "室温匀场"],
            "min_hits": 2,
            "max_terms": 10,
        },
    ],
    "电子学机柜": [
        {
            "name": "电源与接地屏蔽子系统",
            "keywords": ["PDU", "电源", "接地", "屏蔽", "导电", "EMC", "胶条"],
            "min_hits": 2,
            "max_terms": 10,
        },
        {
            "name": "结构与门板安全子系统",
            "keywords": ["横梁", "机柜", "前门", "后下门", "后上门", "钢丝绳", "铰链"],
            "min_hits": 2,
            "max_terms": 10,
        },
        {
            "name": "风道与热设计子系统",
            "keywords": ["风道", "过滤棉", "进风", "排风", "热短路", "CFD", "温升"],
            "min_hits": 2,
            "max_terms": 10,
        },
    ],
}

LIFECYCLE_COVERAGE_PROFILES: list[dict[str, Any]] = [
    {
        "name": "产品类别设计",
        "keywords": ["架构", "设计", "指标", "裕量", "保护", "安全", "功能", "接口", "验证", "竞品"],
        "function_context": "产品设计阶段确保：{function}",
        "effect_context": "设计缺陷会固化到整机基线，造成批量性能、安全或可靠性风险。关联后果：{effect}",
        "cause_context": "需求分解、设计裕量、接口约束或验证覆盖不足。关联机理：{cause}",
        "control_context": "需求评审、架构评审、仿真/样机验证、设计规则检查。现行控制：{controls}",
        "action_context": "补充设计裕量、边界条件验证、保护阈值和评审检查表。建议延伸：{actions}",
    },
    {
        "name": "物流运输",
        "keywords": ["运输", "包装", "振动", "冲击", "跌落", "湿热", "静电", "搬运", "锁止", "到货"],
        "function_context": "物流运输后仍满足：{function}",
        "effect_context": "运输/搬运后隐性损伤会导致到货验收、安装或首次运行失败。关联后果：{effect}",
        "cause_context": "长途振动、冲击、潮湿、ESD、包装固定或运输锁止不足。关联机理：{cause}",
        "control_context": "包装评审、运输试验、冲击/温湿度记录、到货外观与功能检查。现行控制：{controls}",
        "action_context": "增加专用包装、关键器件锁止、运输记录标签、到货复测项目。建议延伸：{actions}",
    },
    {
        "name": "安装调试",
        "keywords": ["安装", "调试", "接线", "校准", "接地", "联调", "验收", "配置", "SOP", "现场"],
        "function_context": "安装调试阶段正确建立：{function}",
        "effect_context": "安装配置偏差会在客户现场暴露为验收失败、性能不稳或安全保护误动作。关联后果：{effect}",
        "cause_context": "接线/接地/配置/校准步骤遗漏，现场条件与工厂验证边界不一致。关联机理：{cause}",
        "control_context": "安装SOP、联调清单、接地/接口核验、出厂参数备份、现场验收测试。现行控制：{controls}",
        "action_context": "增加安装防错、参数模板、自动自检、强制验收记录和现场问题回写。建议延伸：{actions}",
    },
    {
        "name": "客户操作",
        "keywords": ["客户", "操作", "误操作", "权限", "参数", "报警", "提示", "互锁", "培训", "使用"],
        "function_context": "客户日常操作中安全、稳定地完成：{function}",
        "effect_context": "误操作或参数误设会导致实验中断、器件损伤、结果失真或客户投诉。关联后果：{effect}",
        "cause_context": "用户经验差异、参数边界提示不足、权限/互锁不完整或报警解释不清。关联机理：{cause}",
        "control_context": "操作权限、参数范围限制、报警提示、日志追踪、培训材料。现行控制：{controls}",
        "action_context": "增加新手模式、关键操作二次确认、参数推荐、报警指导和操作日志复盘。建议延伸：{actions}",
    },
    {
        "name": "任务执行",
        "keywords": ["任务", "运行", "长时间", "脉冲", "稳定", "精度", "漂移", "动态", "输出", "性能"],
        "function_context": "任务执行期间持续满足：{function}",
        "effect_context": "任务过程中失效会直接影响实验数据、连续运行和核心性能。关联后果：{effect}",
        "cause_context": "长时间运行、动态负载、热漂移、控制滞后或边界工况覆盖不足。关联机理：{cause}",
        "control_context": "在线监测、性能自检、趋势记录、任务前后校验、保护联锁。现行控制：{controls}",
        "action_context": "增加在线诊断、趋势预警、任务前自检、降额策略和异常自动记录。建议延伸：{actions}",
    },
    {
        "name": "环境应力",
        "keywords": ["环境", "温度", "湿度", "EMC", "电磁", "电源", "振动", "粉尘", "冷凝", "散热"],
        "function_context": "环境应力变化下维持：{function}",
        "effect_context": "温湿度、电磁、电源或振动应力会造成间歇性故障、误报警或性能漂移。关联后果：{effect}",
        "cause_context": "客户现场环境超出假设、热/湿/电磁裕量不足或环境监测缺失。关联机理：{cause}",
        "control_context": "环境规范、温湿度/电源/EMC监测、降额设计、环境适应性验证。现行控制：{controls}",
        "action_context": "增加环境传感、EMC/电源裕量、热设计复核、环境超限联锁和客户环境预检。建议延伸：{actions}",
    },
    {
        "name": "维护保养",
        "keywords": ["维护", "保养", "校准", "寿命", "老化", "更换", "备件", "清洁", "诊断", "复位"],
        "function_context": "维护保养后恢复并保持：{function}",
        "effect_context": "维护不到位或寿命件退化会导致重复停机、性能慢性劣化和服务成本上升。关联后果：{effect}",
        "cause_context": "寿命件老化、校准漂移、维护周期不清、备件/清洁/复位策略不足。关联机理：{cause}",
        "control_context": "维护计划、寿命计数、校准记录、备件清单、远程诊断。现行控制：{controls}",
        "action_context": "增加寿命预测、维护提醒、快换设计、远程诊断和保养记录闭环。建议延伸：{actions}",
    },
]

AFMEA_COVERAGE_PROFILES: list[dict[str, Any]] = [
    {
        "name": "存储保管",
        "keywords": ["存储", "仓储", "湿热", "冷凝", "静电", "防尘", "防潮", "长期放置", "保管", "开箱"],
        "target_rows": 4,
        "function_context": "存储保管阶段保持可交付和可安装状态：{function}",
        "effect_context": "存储环境或保管控制失效会在开箱、安装或首次使用时暴露为隐性损伤、性能漂移或客户投诉。关联后果：{effect}",
        "cause_context": "温湿度、冷凝、静电、粉尘、包装防护、保管周期或开箱检查控制不足。关联机理：{cause}",
        "control_context": "仓储环境规范、包装状态检查、温湿度记录、开箱检查、先进先出和保管周期控制。现行控制：{controls}",
        "action_context": "补充仓储环境监控、开箱验收清单、防潮/防静电要求、保管期限和异常复测规则。建议延伸：{actions}",
    },
    *[profile for profile in LIFECYCLE_COVERAGE_PROFILES if profile["name"] != "产品类别设计"],
]

SFMEA_COVERAGE_PROFILES: list[dict[str, Any]] = [
    {
        "name": "系统边界与需求分解",
        "keywords": ["系统", "边界", "需求", "指标", "功能", "架构", "工况", "验收"],
        "target_rows": 5,
        "function_context": "系统边界内正确实现并分解需求：{function}",
        "effect_context": "系统需求或边界失效会造成整机功能缺失、验收失败、性能不稳定或责任归属不清。关联后果：{effect}",
        "cause_context": "需求分解、系统边界、使用工况、接口责任或验收准则定义不足。关联机理：{cause}",
        "control_context": "系统需求评审、边界图、功能链评审、系统验证计划和验收准则。现行控制：{controls}",
        "action_context": "补齐系统边界矩阵、需求追踪、系统验证覆盖和接口责任表。建议延伸：{actions}",
    },
    {
        "name": "子系统接口与集成",
        "keywords": ["子系统", "接口", "连接", "匹配", "集成", "联调", "信号", "机械", "气路", "电源"],
        "target_rows": 5,
        "function_context": "子系统接口在集成状态下正确传递功能：{function}",
        "effect_context": "接口或集成失效会造成系统级功能中断、间歇异常、联调失败或性能降级。关联后果：{effect}",
        "cause_context": "接口协议、机械/电气/流体匹配、边界条件、联调顺序或兼容性定义不足。关联机理：{cause}",
        "control_context": "接口控制文件、集成评审、联调计划、接口测试和系统验收。现行控制：{controls}",
        "action_context": "建立接口矩阵、异常状态联调、边界样机验证和接口变更闭环。建议延伸：{actions}",
    },
    {
        "name": "能量物料信息流",
        "keywords": ["能量", "物料", "信息", "信号", "流量", "热", "电源", "通信", "数据", "控制"],
        "target_rows": 5,
        "function_context": "系统功能链中的能量、物料和信息流保持连续可控：{function}",
        "effect_context": "传递链失效会造成系统输出错误、保护误动作、效率下降或任务中断。关联后果：{effect}",
        "cause_context": "传递路径、容量裕量、状态反馈、同步关系、通信或控制链路定义不足。关联机理：{cause}",
        "control_context": "功能链图、状态监测、通信/功率/流量测试、系统保护和日志。现行控制：{controls}",
        "action_context": "补充传递链裕量验证、状态诊断、链路失效注入和系统级降级策略。建议延伸：{actions}",
    },
    {
        "name": "系统状态与控制逻辑",
        "keywords": ["状态", "控制", "逻辑", "报警", "联锁", "保护", "模式", "配置", "时序"],
        "target_rows": 5,
        "function_context": "系统状态机、保护逻辑和配置管理正确支撑：{function}",
        "effect_context": "系统控制逻辑失效会造成误报警、未保护、错误模式切换、配置错用或安全边界失守。关联后果：{effect}",
        "cause_context": "状态定义、异常转移、保护阈值、配置一致性、时序或诊断策略不足。关联机理：{cause}",
        "control_context": "状态机评审、软件/系统联调、保护阈值验证、异常工况测试和日志审查。现行控制：{controls}",
        "action_context": "增加状态覆盖矩阵、异常注入、配置校验、联锁验证和日志追溯。建议延伸：{actions}",
    },
    {
        "name": "系统环境与外部依赖",
        "keywords": ["环境", "现场", "温度", "湿度", "EMC", "电源", "地线", "客户", "外部"],
        "target_rows": 5,
        "function_context": "系统在外部环境和依赖条件下保持边界功能：{function}",
        "effect_context": "外部依赖或现场边界失效会造成系统性能漂移、误停机、客户验收失败或服务成本上升。关联后果：{effect}",
        "cause_context": "现场电源/接地/温湿度/EMC/空间/操作条件超出系统假设或监测不足。关联机理：{cause}",
        "control_context": "现场条件清单、环境适应性验证、安装验收、系统自检和客户环境预检。现行控制：{controls}",
        "action_context": "补充现场边界要求、安装前检查、环境超限联锁、客户现场数据记录和验收准则。建议延伸：{actions}",
    },
]

DFMEA_PART_DETAIL_PROFILES: list[dict[str, Any]] = [
    {
        "name": "DFMEA零件级-功能架构与接口",
        "keywords": ["接口", "连接", "边界", "信号", "能量", "材料", "信息", "公差", "安装", "匹配"],
        "target_rows": 6,
        "function_context": "零件/接口级实现并传递：{function}",
        "effect_context": "零件级接口或功能链失效会向上游/下游扩散，造成系统性能、安全或可用性风险。关联后果：{effect}",
        "cause_context": "接口定义、公差链、装配基准、接触阻抗、屏蔽/接地、材料兼容或边界条件不足。关联机理：{cause}",
        "control_context": "接口控制文件、BOM/图纸评审、公差链分析、首件/来料/装配验证、系统联调。现行控制：{controls}",
        "action_context": "补齐接口特性矩阵、关键特性 CTQ、装配防错、接口验证和失效注入测试。建议延伸：{actions}",
    },
    {
        "name": "DFMEA零件级-电子器件/PCBA",
        "keywords": ["PCBA", "电阻", "电容", "电感", "MOS", "继电器", "比较器", "ADC", "DAC", "MCU", "光耦", "电源"],
        "target_rows": 6,
        "function_context": "电子器件/PCBA级满足供电、采样、放大、保护、时序或通信要求：{function}",
        "effect_context": "电子器件漂移、开短路、降额不足或焊接缺陷会导致误动作、性能漂移、保护失效或整机停机。关联后果：{effect}",
        "cause_context": "器件降额、热设计、ESD/EOS、焊点疲劳、参数漂移、批次差异、布局串扰或电源完整性不足。关联机理：{cause}",
        "control_context": "原理图/PCB评审、降额清单、DFM/DFT、ICT/FCT、老炼、EMC/ESD和温升验证。现行控制：{controls}",
        "action_context": "补充器件降额、关键节点监测、边界样件验证、失效注入和批次追溯。建议延伸：{actions}",
    },
    {
        "name": "DFMEA零件级-连接件/线束/紧固件/密封件",
        "keywords": ["连接器", "线束", "航插", "端子", "同轴", "BNC", "DB", "螺钉", "紧固", "密封圈", "胶", "屏蔽"],
        "target_rows": 5,
        "function_context": "连接、紧固、屏蔽或密封零件在全寿命内保持：{function}",
        "effect_context": "连接/紧固/密封件失效会造成间歇故障、泄漏、接触不良、EMC退化、松脱或维护返修。关联后果：{effect}",
        "cause_context": "插拔寿命、锁止不足、扭矩窗口、线束应力释放、密封压缩量、胶黏剂老化或误装。关联机理：{cause}",
        "control_context": "选型校核、端子拉力/导通、扭矩标识、防呆、密封测试、插拔寿命和来料检验。现行控制：{controls}",
        "action_context": "增加防松/防错、线束固定、密封压缩验证、插拔寿命和现场可诊断性。建议延伸：{actions}",
    },
]

PFMEA_COVERAGE_PROFILES: list[dict[str, Any]] = [
    {
        "name": "来料与上线准备",
        "keywords": ["来料", "供应商", "上线", "备料", "批次", "检验", "物料", "追溯"],
        "target_rows": 5,
        "function_context": "来料和上线准备过程保证满足工艺输入要求：{function}",
        "effect_context": "来料或备料过程失效会流入装配/测试，导致返工、后工序停滞或客户风险。关联后果：{effect}",
        "cause_context": "供应商批次、来料检验、物料标识、存放、领料或追溯控制不足。关联机理：{cause}",
        "control_context": "IQC、来料检验规范、批次追溯、物料状态标识和上线点检。现行控制：{controls}",
        "action_context": "强化来料抽检、关键特性确认、批次隔离、上线防错和供应商纠正措施。建议延伸：{actions}",
    },
    {
        "name": "装配工序",
        "keywords": ["装配", "拧紧", "定位", "夹具", "工装", "作业", "安装", "扭矩", "顺序"],
        "target_rows": 6,
        "function_context": "装配工序按标准方法形成合格过程输出：{function}",
        "effect_context": "装配过程失效会造成尺寸偏差、松动、错装、漏装、返工或后续功能异常。关联后果：{effect}",
        "cause_context": "作业顺序、夹具定位、扭矩窗口、防错、人员培训或标准作业控制不足。关联机理：{cause}",
        "control_context": "SOP、工装点检、首件确认、扭矩记录、过程巡检和防错确认。现行控制：{controls}",
        "action_context": "增加工装防错、扭矩追溯、作业指导可视化、首件/末件确认和过程审核。建议延伸：{actions}",
    },
    {
        "name": "过程参数与设备状态",
        "keywords": ["参数", "设备", "温度", "压力", "时间", "速度", "校准", "点检", "维护"],
        "target_rows": 6,
        "function_context": "过程参数和设备状态稳定满足工艺窗口：{function}",
        "effect_context": "过程窗口或设备状态失控会造成批量波动、隐性缺陷、返修或检验逃逸。关联后果：{effect}",
        "cause_context": "设备点检、参数设定、校准周期、维护保养、工艺窗口或异常停线规则不足。关联机理：{cause}",
        "control_context": "设备点检、参数锁定、校准记录、SPC、维护计划和异常处理流程。现行控制：{controls}",
        "action_context": "建立参数上下限、SPC预警、设备保养、校准提醒和异常批次隔离。建议延伸：{actions}",
    },
    {
        "name": "检验测试与放行",
        "keywords": ["检验", "测试", "放行", "判定", "治具", "量具", "MSA", "抽检", "全检"],
        "target_rows": 6,
        "function_context": "检验测试过程准确发现并拦截不合格输出：{function}",
        "effect_context": "检验或测试过程失效会造成缺陷流出、误判、重复返工或客户现场失效。关联后果：{effect}",
        "cause_context": "测试覆盖、判定阈值、治具状态、量具能力、MSA、抽样方案或人员判定不足。关联机理：{cause}",
        "control_context": "测试规范、治具点检、量具校准、MSA、抽检/全检规则和放行审批。现行控制：{controls}",
        "action_context": "补充测试覆盖、治具自检、MSA复核、判定规则锁定和缺陷复盘闭环。建议延伸：{actions}",
    },
    {
        "name": "包装入库与交付",
        "keywords": ["包装", "入库", "交付", "标签", "运输", "防护", "出货", "装箱"],
        "target_rows": 5,
        "function_context": "包装入库和交付过程保持产品状态和追溯完整：{function}",
        "effect_context": "包装或交付过程失效会造成运输损伤、错发、漏发、标识错误或客户开箱问题。关联后果：{effect}",
        "cause_context": "包装防护、标签校验、装箱清单、出货检验、运输固定或交付记录不足。关联机理：{cause}",
        "control_context": "包装规范、出货检验、标签扫描、装箱清单、运输防护和交付记录。现行控制：{controls}",
        "action_context": "增加包装验证、扫码防错、装箱复核、运输状态标签和出货抽检闭环。建议延伸：{actions}",
    },
]


@dataclass
class ScopeDefinition:
    name: str
    query_terms: list[str]
    extracted_terms: list[str] = field(default_factory=list)
    auto_suggested: bool = False
    hit_count: int = 0
    reason: str = ""


@dataclass
class DraftRow:
    scope: str
    analysis_object: str
    function: str
    failure_mode: str
    effect: str
    severity: str
    cause: str
    occurrence: str
    current_controls: str
    detection: str
    rpn: str
    recommended_actions: str
    owner: str
    target_date: str
    confirmation_status: str
    rating_basis: str
    reference_type: str
    source_cases: list[str]
    review_comment: str = ""
    post_action_severity: str = ""
    post_action_occurrence: str = ""
    post_action_detection: str = ""
    post_action_rpn: str = ""
    max_match_score: int = 0
    max_scope_hits: int = 0
    confirmation_reasons: list[str] = field(default_factory=list)
    reviewer_focus: str = ""
    boundary_scopes: list[str] = field(default_factory=list)


@dataclass
class InputQualitySignal:
    signal: str
    status: str
    evidence: str
    missing_detail: str = ""


@dataclass
class InputQualityDiagnosis:
    level: str
    summary: str
    signals: list[InputQualitySignal]
    missing_critical_inputs: list[str]
    assumptions: list[str]


@dataclass
class CoverageMatrixItem:
    dimension: str
    status: str
    evidence: str
    review_prompt: str
    reason_tags: list[str] = field(default_factory=list)


@dataclass
class QualityGateFinding:
    gate: str
    status: str
    row_key: str
    finding: str
    required_fix_or_confirmation: str
    reason_tags: list[str] = field(default_factory=list)
    blocking: bool = False


@dataclass
class ConfirmationItem:
    scope: str
    row_key: str
    why_confirmation_is_needed: str
    suggested_reviewer_focus: str
    reference_type: str
    source_cases: list[str]
    review_comment: str = ""
    plain_language_question: str = ""
    why_it_matters: str = ""
    suggested_options: list[str] = field(default_factory=list)
    default_assumption: str = ""
    impact_if_wrong: str = ""
    reason_tags: list[str] = field(default_factory=list)
    priority: str = "medium"
    blocking: bool = False


def load_input_text(args: argparse.Namespace) -> str:
    if args.input_file:
        return Path(args.input_file).read_text(encoding="utf-8").strip()
    if args.input_text:
        return args.input_text.strip()
    raise ValueError("Provide either --input-file or --input-text.")


def valid_token(token: str) -> bool:
    token = token.strip()
    if not token or token in STOPWORDS:
        return False
    if len(token) < 2 or len(token) > 24:
        return False
    if re.fullmatch(r"[\d\-.:%]+", token):
        return False
    if token.lower() in {"rpn", "excel", "json", "csv"}:
        return False
    return True


def build_lexicon() -> list[str]:
    counts: Counter[str] = Counter()
    for json_file in sorted(DATA_ROOT.glob("*/json/*.json")):
        payload = json.loads(json_file.read_text(encoding="utf-8"))
        if payload.get("theme") not in ALLOWED_THEMES:
            continue
        for row in payload.get("filled_rows", []):
            text = row_text(row)
            if not text:
                continue
            for token in tokenize(text):
                if valid_token(token):
                    counts[token] += 1
    ranked = sorted(counts.items(), key=lambda item: (-len(item[0]), -item[1], item[0]))
    return [token for token, _ in ranked]


def extract_query_terms(input_text: str, module: str | None, limit: int = 16) -> list[str]:
    lexicon = build_lexicon()
    found: list[str] = []
    for token in lexicon:
        if token in input_text and token not in found:
            found.append(token)
        if len(found) >= limit:
            break
    if module:
        if module not in found:
            found.insert(0, module)
        for canonical, alias_list in ALIASES.items():
            if module == canonical or module in alias_list:
                for item in [canonical, *alias_list]:
                    if item not in found:
                        found.append(item)
    return found


def parse_scope(raw_scope: str) -> ScopeDefinition:
    if "::" not in raw_scope:
        raise ValueError(f"Invalid --scope value: {raw_scope}. Use 'Scope Name::keyword1 keyword2'.")
    name, raw_terms = raw_scope.split("::", 1)
    terms = [term for term in tokenize(raw_terms) if valid_token(term)]
    if not name.strip() or not terms:
        raise ValueError(f"Invalid --scope value: {raw_scope}.")
    return ScopeDefinition(name=name.strip(), query_terms=terms, auto_suggested=False, reason="manual")


def canonicalize_module(module: str) -> str:
    if module in SCOPE_PROFILES:
        return module
    for canonical, aliases in ALIASES.items():
        if module == canonical or module in aliases:
            return canonical
    return module


def module_family_rank(sheet: str, module: str | None) -> int:
    canonical_module = canonicalize_module(module) if module else None
    if not canonical_module:
        return 0

    same_names = {canonical_module, *ALIASES.get(canonical_module, [])}
    if sheet in same_names:
        return 2

    related_modules = RELATED_MODULES.get(canonical_module, [])
    related_names = {name for related in related_modules for name in [related, *ALIASES.get(related, [])]}
    if sheet in related_names:
        return 1

    return 0


def suggest_scopes(module: str, input_text: str, extracted_terms: list[str]) -> list[ScopeDefinition]:
    canonical_module = canonicalize_module(module)
    profiles = SCOPE_PROFILES.get(canonical_module, [])
    lowered = input_text.lower()
    suggestions: list[ScopeDefinition] = []

    for profile in profiles:
        keywords = profile["keywords"]
        hit_keywords = [keyword for keyword in keywords if keyword.lower() in lowered]
        if len(hit_keywords) < profile["min_hits"]:
            continue

        query_terms: list[str] = []
        for keyword in hit_keywords:
            if keyword not in query_terms:
                query_terms.append(keyword)
        for term in extracted_terms:
            if term in keywords and term not in query_terms:
                query_terms.append(term)
        for keyword in keywords:
            if keyword not in query_terms:
                query_terms.append(keyword)
            if len(query_terms) >= profile["max_terms"]:
                break

        suggestions.append(
            ScopeDefinition(
                name=profile["name"],
                query_terms=query_terms[: profile["max_terms"]],
                extracted_terms=[term for term in extracted_terms if term in query_terms],
                auto_suggested=True,
                hit_count=len(hit_keywords),
                reason=f"matched keywords: {' / '.join(hit_keywords[:8])}",
            )
        )

    suggestions.sort(key=lambda scope: (-scope.hit_count, scope.name))
    return suggestions


def normalize_fmea_type(fmea_type: str) -> str:
    normalized = (fmea_type or "DFMEA").upper()
    if normalized not in VALID_FMEA_TYPES:
        raise ValueError(f"Unsupported FMEA type: {fmea_type}. Use one of {', '.join(sorted(VALID_FMEA_TYPES))}.")
    return normalized


def coverage_profiles_for(fmea_type: str, coverage_mode: str) -> list[dict[str, Any]]:
    normalized_type = normalize_fmea_type(fmea_type)
    if normalized_type == "AFMEA":
        return AFMEA_COVERAGE_PROFILES
    if normalized_type == "SFMEA":
        return SFMEA_COVERAGE_PROFILES
    if normalized_type == "PFMEA":
        return PFMEA_COVERAGE_PROFILES
    if coverage_mode == "part":
        return DFMEA_PART_DETAIL_PROFILES
    return LIFECYCLE_COVERAGE_PROFILES


def template_path_for(fmea_type: str) -> Path:
    normalized_type = normalize_fmea_type(fmea_type)
    type_template = TEMPLATE_PATHS_BY_TYPE.get(normalized_type)
    if type_template and type_template.exists():
        return type_template
    return DEFAULT_TEMPLATE_PATH


def suggest_lifecycle_scopes(
    module: str,
    input_text: str,
    extracted_terms: list[str],
    profiles: list[dict[str, Any]] | None = None,
) -> list[ScopeDefinition]:
    lowered = input_text.lower()
    scopes: list[ScopeDefinition] = []
    for profile in profiles or LIFECYCLE_COVERAGE_PROFILES:
        keywords = profile["keywords"]
        hit_keywords = [keyword for keyword in keywords if keyword.lower() in lowered]
        query_terms: list[str] = []
        for term in [module, *hit_keywords, *extracted_terms[:10], *keywords]:
            if term and term not in query_terms:
                query_terms.append(term)
        scopes.append(
            ScopeDefinition(
                name=profile["name"],
                query_terms=query_terms[:18],
                extracted_terms=[term for term in extracted_terms if term in query_terms],
                auto_suggested=True,
                hit_count=len(hit_keywords),
                reason=(
                    f"type coverage profile; matched keywords: {' / '.join(hit_keywords[:8])}"
                    if hit_keywords
                    else "type coverage profile; added to avoid narrow subsystem-only FMEA coverage"
                ),
            )
        )
    return scopes


def normalize_space(value: str) -> str:
    return re.sub(r"\s+", " ", value).strip()


def read_match_row(match: Match, cache: dict[Path, Any]) -> dict[str, str]:
    json_path = PROJECT_ROOT / match.source
    if json_path not in cache:
        cache[json_path] = json.loads(json_path.read_text(encoding="utf-8"))
    payload = cache[json_path]
    for row in payload.get("filled_rows", []):
        if row.get("__excel_row__") == match.excel_row:
            return row
    raise KeyError(f"Could not find row {match.excel_row} in {json_path}")


def first_value(row: dict[str, str], field_name: str) -> str:
    for key in FIELD_ALIASES[field_name]:
        value = normalize_space(row.get(key, ""))
        if value:
            return value
    return ""


def safe_int(value: str) -> int | None:
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return None


def compute_rpn(severity: str, occurrence: str, detection: str, current_rpn: str) -> str:
    if current_rpn:
        return current_rpn
    s_value = safe_int(severity)
    o_value = safe_int(occurrence)
    d_value = safe_int(detection)
    if s_value is None or o_value is None or d_value is None:
        return ""
    return str(s_value * o_value * d_value)


def compute_post_action_rpn(severity: str, occurrence: str, detection: str, current_rpn: str) -> str:
    return compute_rpn(severity, occurrence, detection, current_rpn)


def fill_missing_draft_scores(
    severity: str,
    occurrence: str,
    detection: str,
    rpn: str = "",
) -> tuple[str, str, str, str, list[str]]:
    """Fill missing S/O/D with conservative AI draft values so rows remain reviewable.

    Missing source scores are still routed to the confirmation queue by callers.
    These defaults are not enterprise facts; they are placeholders that make RPN
    sorting and workbook formulas usable before expert calibration.
    """

    missing = [label for label, value in [("S", severity), ("O", occurrence), ("D", detection)] if not value]
    severity = severity or "7"
    occurrence = occurrence or "5"
    detection = detection or "5"
    rpn = compute_rpn(severity, occurrence, detection, rpn)
    return severity, occurrence, detection, rpn, missing


def combine_unique(values: list[str]) -> str:
    seen: list[str] = []
    for value in values:
        cleaned = normalize_space(value)
        if cleaned and cleaned not in seen:
            seen.append(cleaned)
    return "\n".join(seen)


def dedupe_key(row: dict[str, str]) -> tuple[str, str, str, str]:
    return (
        first_value(row, "analysis_object"),
        first_value(row, "function"),
        first_value(row, "failure_mode"),
        first_value(row, "effect"),
    )


def scope_hit_count(row: dict[str, str], scope_terms: list[str]) -> int:
    primary_text = " ".join(
        [
            first_value(row, "analysis_object"),
            first_value(row, "function"),
            first_value(row, "failure_mode"),
            first_value(row, "effect"),
            first_value(row, "cause"),
        ]
    )
    lowered = primary_text.lower()
    return sum(1 for term in scope_terms if term.lower() in lowered)


def scope_focus_score(row: dict[str, str], scope_terms: list[str]) -> int:
    focus_text = " ".join(
        [
            first_value(row, "analysis_object"),
            first_value(row, "function"),
            first_value(row, "failure_mode"),
        ]
    ).lower()
    context_text = " ".join(
        [
            first_value(row, "effect"),
            first_value(row, "cause"),
        ]
    ).lower()
    score = 0
    for term in scope_terms:
        lowered = term.lower()
        if lowered in focus_text:
            score += 2
        elif lowered in context_text:
            score += 1
    return score


def best_scope_name(row: dict[str, str], scopes: list[ScopeDefinition]) -> str | None:
    ranked = []
    for index, scope in enumerate(scopes):
        ranked.append(
            (
                scope_focus_score(row, scope.query_terms),
                scope_hit_count(row, scope.query_terms),
                -index,
                scope.name,
            )
        )
    best = max(ranked)
    if best[0] <= 0 and best[1] <= 0:
        return None
    return best[3]


def boundary_scope_names(row: dict[str, str], current_scope: str, scopes: list[ScopeDefinition]) -> list[str]:
    current_focus = 0
    current_hits = 0
    candidates: list[tuple[int, int, str]] = []

    for scope in scopes:
        focus = scope_focus_score(row, scope.query_terms)
        hits = scope_hit_count(row, scope.query_terms)
        if scope.name == current_scope:
            current_focus = focus
            current_hits = hits
        if focus > 0 or hits > 0:
            candidates.append((focus, hits, scope.name))

    boundaries: list[str] = []
    for focus, hits, scope_name in candidates:
        if scope_name == current_scope:
            continue
        if current_focus > 0:
            if focus >= current_focus - 1 and hits > 0:
                boundaries.append(scope_name)
        elif current_hits > 0 and hits >= current_hits:
            boundaries.append(scope_name)

    return boundaries


def build_reference_type(matches_for_row: list[Match], module: str | None) -> str:
    family_ranks = [module_family_rank(match.sheet, module) for match in matches_for_row]
    if any(rank >= 2 for rank in family_ranks):
        return "current module"
    if any(rank == 1 for rank in family_ranks):
        return "direct family reference"
    return "broader analogy"


def build_confirmation_reasons(
    reference_type: str,
    boundary_scopes: list[str],
    themes: set[str],
    severity: str,
    occurrence: str,
    detection: str,
) -> list[str]:
    reasons: list[str] = []
    missing_scores = [label for label, value in [("S", severity), ("O", occurrence), ("D", detection)] if not value]
    if missing_scores:
        reasons.append(f"源案例缺少完整评分字段：{'/'.join(missing_scores)}")
    elif not occurrence or not detection:
        reasons.append("O/D 评分依据仍不完整")

    if reference_type != "current module":
        reasons.append("该行主要来自跨模块家族类比，需要确认是否适用于当前模块责任范围")

    if boundary_scopes:
        reasons.append(f"scope 归属存在边界，当前放在本 scope，但也可能属于：{' / '.join(boundary_scopes)}")

    if any(theme != "dfmea_sample_data" for theme in themes):
        reasons.append("来源中包含模板或知识库条目，需要确认是否已落到具体机理和现行控制")

    return reasons


def build_confirmation_status(reasons: list[str]) -> str:
    if reasons:
        return "needs expert confirmation"
    return "draft"


def build_rating_basis(
    themes: set[str],
    reference_type: str,
    severity: str,
    occurrence: str,
    detection: str,
) -> str:
    parts: list[str] = []

    if severity:
        parts.append(f"S={severity} 继承自历史案例的后果强度")
    else:
        parts.append("S 未在源案例中明确给出")

    if occurrence and detection:
        if themes == {"dfmea_sample_data"}:
            parts.append(f"O={occurrence}、D={detection} 继承自成品 DFMEA，但仍需结合本企业标尺校准")
        else:
            parts.append(f"O={occurrence}、D={detection} 来自混合来源，需结合真实试验与控制能力校准")
    else:
        parts.append("O/D 未完整继承，需专家补齐评分依据")

    if reference_type != "current module":
        parts.append(f"本行属于 {reference_type}，仅建议作为补缺参考")

    return "；".join(parts)


def build_reviewer_focus(reference_type: str, boundary_scopes: list[str], occurrence: str, detection: str) -> str:
    focus_points: list[str] = []
    if boundary_scopes:
        focus_points.append("确认 scope 归属与责任边界")
    if reference_type != "current module":
        focus_points.append("确认该类比是否适用于当前模块机理、接口和责任范围")
    if not occurrence or not detection:
        focus_points.append("补齐 O/D 与现行检测控制")
    else:
        focus_points.append("校准 O/D 与实际保护、测试和筛选能力是否匹配")
    return "；".join(focus_points)


INPUT_QUALITY_RULES: list[dict[str, Any]] = [
    {
        "signal": "module_or_object",
        "label": "模块或分析对象",
        "keywords": [],
        "missing": "确认具体模块、子系统或零件边界",
        "assumption": "以命令行模块名作为分析对象",
    },
    {
        "signal": "function_or_requirement",
        "label": "关键功能或要求",
        "keywords": ["功能", "用于", "实现", "要求", "指标", "性能", "放大", "检测", "控制", "保护", "传输", "支撑", "冷却", "供电"],
        "missing": "补充该对象必须实现的核心功能、性能指标或设计要求",
        "assumption": "按模块名称和历史案例推断功能",
    },
    {
        "signal": "scenario_or_lifecycle",
        "label": "使用场景或生命周期",
        "keywords": ["贮存", "储存", "运输", "安装", "调试", "使用", "运行", "操作", "维护", "保养", "客户", "现场", "任务", "移机"],
        "missing": "补充贮存、运输、安装、运行、维护等实际场景",
        "assumption": "按通用生命周期补齐场景",
    },
    {
        "signal": "environment",
        "label": "环境和应用应力",
        "keywords": ["温度", "湿度", "振动", "冲击", "EMC", "电磁", "ESD", "粉尘", "冷凝", "散热", "海拔", "电源", "噪声"],
        "missing": "补充温湿度、振动、EMC、电源、散热或污染等边界条件",
        "assumption": "按同类设备常见环境应力推断",
    },
    {
        "signal": "interfaces",
        "label": "接口关系",
        "keywords": ["接口", "连接", "线缆", "接插件", "信号", "能量", "气路", "液路", "运动", "通信", "CAN", "RS", "以太网", "控制"],
        "missing": "补充结构、信号、能量、流体、运动或控制接口",
        "assumption": "从相似模块案例推断接口风险",
    },
    {
        "signal": "bom_or_key_parts",
        "label": "BOM/关键件/材料",
        "keywords": ["BOM", "物料", "零件", "器件", "材料", "电阻", "电容", "传感器", "继电器", "风扇", "线圈", "接头", "板卡"],
        "missing": "补充关键零部件、材料、连接方式或特殊工艺",
        "assumption": "按模块历史案例中的典型关键件推断",
    },
    {
        "signal": "current_controls",
        "label": "现行控制/测试/报警",
        "keywords": ["测试", "验证", "检验", "检查", "报警", "联锁", "保护", "监测", "自检", "筛选", "评审", "SOP"],
        "missing": "补充现有预防控制、探测控制、测试、报警或联锁",
        "assumption": "以通用测试和保护措施作为草稿控制",
    },
    {
        "signal": "historical_issues",
        "label": "历史问题或相似案例",
        "keywords": ["历史", "故障", "失效", "投诉", "维修", "返修", "案例", "问题", "异常", "经验", "教训"],
        "missing": "补充历史故障、维修、投诉或相似 FMEA 行",
        "assumption": "仅引用资料库相似案例，不代表当前模块已发生",
    },
    {
        "signal": "impact_context",
        "label": "客户/后工序影响",
        "keywords": ["客户", "后工序", "安全", "停机", "损坏", "性能", "精度", "投诉", "验收", "返工", "法规", "合规"],
        "missing": "补充对客户任务、后工序、安全、成本或合规的影响",
        "assumption": "按核心功能受影响估计严重度",
    },
    {
        "signal": "scoring_evidence",
        "label": "S/O/D 评分证据",
        "keywords": ["S", "O", "D", "RPN", "频度", "严重度", "探测度", "发生", "检出", "过程能力", "良率", "ppm", "覆盖率"],
        "missing": "补充企业评分标尺、发生频度、测试覆盖或检出能力证据",
        "assumption": "O/D 只作为 AI 草稿，必须校准",
    },
]

COVERAGE_DIMENSIONS: dict[str, list[dict[str, Any]]] = {
    "AFMEA": [
        {"dimension": "贮存", "keywords": ["贮存", "储存", "仓储", "温湿度", "防潮"]},
        {"dimension": "物流运输", "keywords": ["运输", "包装", "振动", "冲击", "跌落", "搬运"]},
        {"dimension": "安装调试", "keywords": ["安装", "调试", "接线", "校准", "验收"]},
        {"dimension": "正常操作", "keywords": ["操作", "使用", "运行", "任务", "客户"]},
        {"dimension": "异常/误操作", "keywords": ["误操作", "异常", "报警", "权限", "互锁"]},
        {"dimension": "维护保养", "keywords": ["维护", "保养", "校准", "寿命", "备件"]},
        {"dimension": "移机/场地变化", "keywords": ["移机", "场地", "搬迁", "重新安装"]},
        {"dimension": "报废/退役", "keywords": ["报废", "退役", "处置", "回收"]},
    ],
    "SFMEA": [
        {"dimension": "系统分解", "keywords": ["系统", "子系统", "模块", "架构", "分解"]},
        {"dimension": "子系统功能", "keywords": ["功能", "要求", "指标", "任务"]},
        {"dimension": "结构接口", "keywords": ["结构", "连接", "安装", "固定", "装配"]},
        {"dimension": "信号/数据接口", "keywords": ["信号", "通信", "数据", "CAN", "RS", "接口"]},
        {"dimension": "能量/物料接口", "keywords": ["供电", "能量", "气路", "液路", "冷媒", "射频"]},
        {"dimension": "边界 ownership", "keywords": ["边界", "责任", "归属", "接口", "联调"]},
    ],
    "DFMEA": [
        {"dimension": "功能/要求覆盖", "keywords": ["功能", "要求", "指标", "性能"]},
        {"dimension": "关键件/部件覆盖", "keywords": ["零件", "部件", "器件", "BOM", "物料", "材料"]},
        {"dimension": "失效原因机理", "keywords": ["原因", "机理", "老化", "疲劳", "漂移", "短路", "开路", "磨损"]},
        {"dimension": "设计约束/裕量", "keywords": ["裕量", "容差", "约束", "阈值", "散热", "强度", "EMC"]},
        {"dimension": "供应商/制造影响", "keywords": ["供应商", "制造", "装配", "焊接", "来料", "过程", "一致性"]},
        {"dimension": "现行预防/探测控制", "keywords": ["测试", "验证", "检验", "报警", "联锁", "保护", "监测"]},
    ],
    "PFMEA": [
        {"dimension": "来料与上线准备", "keywords": ["来料", "供应商", "上线", "备料", "批次", "追溯"]},
        {"dimension": "装配工序", "keywords": ["装配", "拧紧", "定位", "夹具", "工装", "扭矩", "作业"]},
        {"dimension": "过程参数与设备状态", "keywords": ["参数", "设备", "温度", "压力", "时间", "速度", "校准", "点检"]},
        {"dimension": "检验测试与放行", "keywords": ["检验", "测试", "放行", "判定", "治具", "量具", "MSA"]},
        {"dimension": "包装入库与交付", "keywords": ["包装", "入库", "交付", "标签", "运输", "防护", "出货"]},
        {"dimension": "后工序与客户影响", "keywords": ["后工序", "客户", "流出", "返工", "联调", "投诉"]},
    ],
}

VAGUE_ACTION_PATTERNS = [
    "加强培训",
    "加强检查",
    "优化设计",
    "提高质量",
    "图纸审核",
    "采购认证",
    "定期检查",
    "规范操作",
    "加强管理",
    "改进管理",
    "持续改进",
]

ACTION_SPECIFIC_KEYWORDS = [
    "防错",
    "防呆",
    "互锁",
    "限位",
    "夹具",
    "工装",
    "治具",
    "定扭矩",
    "扭矩",
    "扫码",
    "条码",
    "二维码",
    "传感器",
    "报警",
    "阈值",
    "SPC",
    "MSA",
    "GR&R",
    "点检",
    "校准",
    "追溯",
    "冗余",
    "失效注入",
    "硬线",
    "AND",
    "参数锁定",
    "控制计划",
    "首件",
    "末件",
    "全检",
    "抽检",
    "不对称",
    "防反",
    "BOM",
    "材料锁定",
    "降额",
    "测试覆盖",
    "FCT",
    "ICT",
    "上下限",
    "控制图",
    "维护计划",
    "隔离",
    "放行",
    "自动",
    "记录",
    "拍照",
]

PHYSICS_DOMAIN_KEYWORDS = [
    "磨损",
    "摩擦",
    "间隙",
    "刚度",
    "变形",
    "松动",
    "干涉",
    "扭矩",
    "压力",
    "泄漏",
    "流量",
    "密封",
    "冷凝",
    "污染",
    "真空",
    "开路",
    "短路",
    "漂移",
    "噪声",
    "接地",
    "屏蔽",
    "降额",
    "时序",
    "通信",
    "状态",
    "阈值",
    "报警",
    "联锁",
    "配置",
    "日志",
    "看门狗",
    "热",
    "温度",
    "老化",
    "脆化",
    "膨胀",
    "腐蚀",
    "兼容",
    "疲劳",
    "工位",
    "工装",
    "夹具",
    "治具",
    "参数",
    "MSA",
    "SPC",
    "检验",
    "放行",
]

GENERIC_CAUSE_PATTERNS = [
    "设计不足",
    "控制不足",
    "管理不足",
    "考虑不足",
    "验证不足",
    "缺少本维度历史案例",
    "风险未充分识别",
    "异常",
    "问题",
    "不足",
]

MANUAL_ONLY_DETECTION_KEYWORDS = ["人工", "目视", "检查", "评审", "审核", "巡检"]

STRONG_DETECTION_KEYWORDS = [
    "自动",
    "互锁",
    "传感器",
    "报警",
    "FCT",
    "ICT",
    "全检",
    "扫码",
    "SPC",
    "MSA",
    "治具自检",
    "在线",
    "记录",
    "追溯",
]

IMPACT_KEYWORDS = [
    "客户",
    "后工序",
    "安全",
    "法规",
    "停机",
    "损坏",
    "性能",
    "精度",
    "验收",
    "投诉",
    "返工",
    "流出",
    "交付",
    "服务",
]

TYPE_BOUNDARY_REVIEW_TERMS: dict[str, dict[str, Any]] = {
    "AFMEA": {
        "wrong_terms": ["BOM", "材料", "材质", "公差", "降额", "PCBA", "焊点", "焊接", "图纸"],
        "expected_terms": ["客户", "操作", "使用", "安装", "运输", "维护", "现场", "断电", "断气", "误操作", "生命周期"],
        "message": "AFMEA 应以应用场景、操作流和突发事件为主,当前行疑似落到了设计/BOM/材料问题。",
    },
    "SFMEA": {
        "wrong_terms": ["BOM", "材料", "材质", "公差", "工序", "操作员", "作业员", "师傅", "来料", "放行"],
        "expected_terms": ["接口", "边界", "系统", "子系统", "通信", "时序", "能量", "物料", "信息", "状态", "联调"],
        "message": "SFMEA 应聚焦系统边界、接口、功能链和状态逻辑,当前行疑似落到了零件或过程问题。",
    },
    "DFMEA": {
        "wrong_terms": ["工序", "工位", "操作员", "作业员", "师傅", "首件", "末件", "放行", "包装入库", "过程巡检"],
        "expected_terms": ["设计", "BOM", "材料", "材质", "公差", "降额", "热", "EMC", "PCBA", "连接器", "图纸", "裕量"],
        "message": "DFMEA 应聚焦设计对象、材料、选型、裕量和物理极限,当前行疑似落到了制造过程控制。",
    },
    "PFMEA": {
        "wrong_terms": ["设计裕量", "降额不足", "材料不耐", "材质不耐", "选型错误", "图纸公差", "热设计不足", "EMC裕量"],
        "expected_terms": ["工序", "工位", "装配", "测试", "检验", "工装", "夹具", "治具", "设备", "扭矩", "参数", "放行"],
        "message": "PFMEA 应聚焦工序、工装、设备、检验和放行控制,当前行疑似落到了设计机理本身。",
    },
}


def keyword_hit(text: str, keywords: list[str]) -> str:
    lowered = text.lower()
    for keyword in keywords:
        if keyword and keyword.lower() in lowered:
            return keyword
    return ""


def diagnose_input_quality(module: str, fmea_type: str, input_text: str, scopes: list[ScopeDefinition]) -> InputQualityDiagnosis:
    combined = " ".join([module, fmea_type, input_text])
    signals: list[InputQualitySignal] = []
    missing: list[str] = []
    assumptions: list[str] = []

    for rule in INPUT_QUALITY_RULES:
        if rule["signal"] == "module_or_object":
            has_signal = bool(module or keyword_hit(combined, ["模块", "系统", "零件", "部件", "产品"]))
            evidence = module or "输入文本包含分析对象线索"
        else:
            hit = keyword_hit(combined, rule["keywords"])
            has_signal = bool(hit)
            evidence = f"命中关键词：{hit}" if hit else ""

        if has_signal:
            status = "present"
            missing_detail = ""
        else:
            status = "missing"
            evidence = "未在用户输入中识别到"
            missing_detail = rule["missing"]
            missing.append(rule["missing"])
            assumptions.append(rule["assumption"])

        signals.append(
            InputQualitySignal(
                signal=rule["signal"],
                status=status,
                evidence=evidence,
                missing_detail=missing_detail,
            )
        )

    missing_count = len(missing)
    if missing_count <= 2:
        level = "strong"
        summary = "输入已覆盖主要 FMEA 起草要素，仍需按企业标尺校准 O/D。"
    elif missing_count <= 5:
        level = "usable_with_assumptions"
        summary = "输入可支持首版草稿，但若干关键信息需要在评审中确认。"
    else:
        level = "high_risk_missing_context"
        summary = "输入偏薄，草稿只能作为发现风险和引导补充信息的起点。"

    return InputQualityDiagnosis(
        level=level,
        summary=summary,
        signals=signals,
        missing_critical_inputs=missing[:6],
        assumptions=assumptions[:6],
    )


def row_search_text(row: DraftRow) -> str:
    return " ".join(
        [
            row.scope,
            row.analysis_object,
            row.function,
            row.failure_mode,
            row.effect,
            row.cause,
            row.current_controls,
            row.recommended_actions,
            row.rating_basis,
        ]
    )


def build_coverage_matrix(
    fmea_type: str,
    scopes: list[ScopeDefinition],
    scope_rows: dict[str, list[DraftRow]],
    input_quality: InputQualityDiagnosis,
) -> list[CoverageMatrixItem]:
    dimensions = COVERAGE_DIMENSIONS.get(fmea_type.upper(), COVERAGE_DIMENSIONS["DFMEA"])
    rows = [row for group in scope_rows.values() for row in group]
    scope_text = " ".join(scope.name + " " + " ".join(scope.query_terms) for scope in scopes)
    missing_signals = {signal.signal for signal in input_quality.signals if signal.status == "missing"}
    matrix: list[CoverageMatrixItem] = []

    for item in dimensions:
        dimension = item["dimension"]
        keywords = item["keywords"]
        matched_rows = [row for row in rows if keyword_hit(row_search_text(row), keywords)]
        scope_hit = keyword_hit(scope_text, keywords)
        if matched_rows:
            weak_reasons: list[str] = []
            if all(row.reference_type != "current module" for row in matched_rows):
                weak_reasons.append("主要依赖类比来源")
            if any(row.confirmation_status == "needs expert confirmation" for row in matched_rows):
                weak_reasons.append("存在待确认评分或适用性")
            if any(not normalize_space(row.current_controls) for row in matched_rows):
                weak_reasons.append("现行控制不充分")
            status = "weak" if weak_reasons else "covered"
            evidence = f"关联行数：{len(matched_rows)}" + (f"；{'；'.join(weak_reasons)}" if weak_reasons else "")
        elif scope_hit:
            status = "weak"
            evidence = f"scope 或输入命中关键词：{scope_hit}，但草稿行支撑不足"
        else:
            status = "missing"
            evidence = "未识别到对应草稿行或 scope 线索"

        tags = ["coverage_gap"] if status in {"weak", "missing"} else []
        if status == "missing":
            tags.append("missing_dimension")
        if "current_controls" in missing_signals and dimension in {"现行预防/探测控制", "正常操作", "子系统功能"}:
            tags.append("missing_controls")
        prompt = f"确认 `{dimension}` 是否适用于当前分析范围；如适用，请补充场景、控制措施或历史问题。"
        matrix.append(CoverageMatrixItem(dimension=dimension, status=status, evidence=evidence, review_prompt=prompt, reason_tags=tags))

    return matrix


def has_numbered_structure(text: str) -> bool:
    return bool(re.search(r"(^|\n)\s*\d+[.、]\s*", text))


def add_quality_finding(
    findings: list[QualityGateFinding],
    gate: str,
    status: str,
    row_key_value: str,
    finding: str,
    required_fix_or_confirmation: str,
    reason_tags: list[str],
    blocking: bool = False,
) -> None:
    candidate = QualityGateFinding(
        gate=gate,
        status=status,
        row_key=row_key_value,
        finding=finding,
        required_fix_or_confirmation=required_fix_or_confirmation,
        reason_tags=reason_tags,
        blocking=blocking,
    )
    fingerprint = (candidate.gate, candidate.row_key, candidate.finding)
    if any((item.gate, item.row_key, item.finding) == fingerprint for item in findings):
        return
    findings.append(candidate)


def build_quality_gate_findings(fmea_type: str, scope_rows: dict[str, list[DraftRow]]) -> list[QualityGateFinding]:
    normalized_type = normalize_fmea_type(fmea_type)
    rows = [row for group in scope_rows.values() for row in group]
    findings: list[QualityGateFinding] = []
    boundary_rule = TYPE_BOUNDARY_REVIEW_TERMS[normalized_type]

    for row in rows:
        key = row_key(row)
        cause_text = normalize_space(row.cause)
        action_text = normalize_space(row.recommended_actions)
        effect_text = normalize_space(row.effect)
        primary_text = normalize_space(" ".join([row.analysis_object, row.function, row.failure_mode, row.cause]))
        full_text = normalize_space(row_search_text(row))

        wrong_hit = keyword_hit(primary_text, boundary_rule["wrong_terms"])
        expected_hit = keyword_hit(primary_text, boundary_rule["expected_terms"])
        if wrong_hit and not expected_hit:
            add_quality_finding(
                findings,
                "type_boundary",
                "review",
                key,
                f"命中疑似串台关键词 `{wrong_hit}`。{boundary_rule['message']}",
                "确认该行是否应改写到当前 FMEA 类型的分析对象；若属于另一类 FMEA,拆分或迁移。",
                ["quality_gate", "fmea_type_boundary", normalized_type.lower()],
                blocking=True,
            )

        cause_is_too_generic = (
            not cause_text
            or len(cause_text) < 8
            or any(pattern in cause_text for pattern in GENERIC_CAUSE_PATTERNS)
        )
        has_physics_domain = bool(keyword_hit(primary_text, PHYSICS_DOMAIN_KEYWORDS))
        if cause_is_too_generic or not has_physics_domain:
            add_quality_finding(
                findings,
                "physics_self_consistency",
                "review",
                key,
                "失效原因/机理偏泛,尚未体现清晰的物理、接口、控制或过程机制。",
                "把原因改写为对象域内的具体机制,并确认 failure mode -> effect -> action 能闭环。",
                ["quality_gate", "weak_physics"],
                blocking=cause_is_too_generic,
            )

        if effect_text and not keyword_hit(effect_text, IMPACT_KEYWORDS):
            add_quality_finding(
                findings,
                "physics_self_consistency",
                "pass_with_note",
                key,
                "失效影响未明确客户、系统、安全、后工序、交付或服务后果。",
                "补充影响对象和后果链,避免只有功能性描述。",
                ["quality_gate", "weak_effect_chain"],
                blocking=False,
            )

        detection_value = safe_int(row.detection)
        controls_text = normalize_space(row.current_controls)
        if detection_value is not None and detection_value <= 3:
            manual_hit = keyword_hit(controls_text, MANUAL_ONLY_DETECTION_KEYWORDS)
            strong_hit = keyword_hit(controls_text, STRONG_DETECTION_KEYWORDS)
            if manual_hit and not strong_hit:
                add_quality_finding(
                    findings,
                    "physics_self_consistency",
                    "review",
                    key,
                    f"D={row.detection} 但现行控制主要是 `{manual_hit}` 类人工控制,检出能力可能被高估。",
                    "确认是否存在自动测试、互锁、传感器、全检、治具自检或数据追溯；否则重估 D。",
                    ["quality_gate", "detection_score_mismatch"],
                    blocking=True,
                )

        vague_hit = keyword_hit(action_text, VAGUE_ACTION_PATTERNS)
        has_specific_action = bool(keyword_hit(action_text, ACTION_SPECIFIC_KEYWORDS) or re.search(r"\d", action_text))
        if not action_text:
            add_quality_finding(
                findings,
                "actionability",
                "fail",
                key,
                "建议措施为空,无法进入 OpenClaw 评审或责任闭环。",
                "补充可执行措施,至少说明设计/工艺/测试/控制计划动作、责任角色和验收证据。",
                ["quality_gate", "missing_action", "poka_yoke_missing"],
                blocking=True,
            )
        elif vague_hit and not has_specific_action:
            add_quality_finding(
                findings,
                "actionability",
                "fail",
                key,
                f"建议措施包含 `{vague_hit}` 这类泛化动作,但没有可量化或可落地的控制。",
                "改写为能进入 BOM、图纸、工装、SOP、控制计划、测试计划或 OpenClaw 行动卡的措施。",
                ["quality_gate", "vague_action", "poka_yoke_missing"],
                blocking=True,
            )
        elif not has_specific_action and len(action_text) < 24:
            add_quality_finding(
                findings,
                "actionability",
                "review",
                key,
                "建议措施过短,防错/探测/责任闭环不清晰。",
                "补充具体防错、测试、控制计划、记录追溯或验收条件。",
                ["quality_gate", "weak_actionability", "poka_yoke_missing"],
                blocking=False,
            )

        for field_name, value in [
            ("失效影响", row.effect),
            ("失效原因", row.cause),
            ("现行控制", row.current_controls),
            ("建议措施", row.recommended_actions),
        ]:
            normalized_value = normalize_space(value)
            if len(normalized_value) >= 90 and not has_numbered_structure(value):
                add_quality_finding(
                    findings,
                    "formatting",
                    "pass_with_note",
                    key,
                    f"{field_name} 内容较长但未使用 `1.`/`2.`/`3.` 分层编号。",
                    "将长单元格拆成多级编号,保持原因、影响、控制和措施的逻辑对应。",
                    ["quality_gate", "formatting_gate", "long_cell_structure"],
                    blocking=False,
                )

        if normalized_type == "PFMEA" and not keyword_hit(full_text, ["工序", "工位", "装配", "测试", "检验", "工装", "夹具", "治具", "设备", "参数", "放行"]):
            add_quality_finding(
                findings,
                "type_boundary",
                "review",
                key,
                "PFMEA 行未明显落到工序、设备工装、过程参数、检验测试或放行控制。",
                "确认该行是否应改写为过程失效,或迁移到 DFMEA/SFMEA。",
                ["quality_gate", "fmea_type_boundary", "pfmea_process_focus"],
                blocking=True,
            )

    return findings


def build_quality_gate_confirmation_items(quality_gate_findings: list[QualityGateFinding]) -> list[ConfirmationItem]:
    items: list[ConfirmationItem] = []
    for finding in quality_gate_findings:
        if not finding.blocking and finding.status != "fail":
            continue
        tags = [*finding.reason_tags, "non_expert_validation"]
        items.append(
            ConfirmationItem(
                scope="质量门禁",
                row_key=finding.row_key,
                why_confirmation_is_needed=f"{finding.gate}: {finding.finding}",
                suggested_reviewer_focus=finding.required_fix_or_confirmation,
                reference_type="AI quality gate",
                source_cases=[],
                plain_language_question=f"这条 `{finding.row_key}` 是否应按质量门禁修正,或由对应专家确认后保留?",
                why_it_matters="质量门禁问题可能改变 FMEA 类型归属、物理机理、D/O 评分或措施优先级。",
                suggested_options=["按门禁修正", "专家确认后保留", "迁移到其他 FMEA 类型", "删除该行"],
                default_assumption="先作为阻塞评审项保留",
                impact_if_wrong="可能把串台、物理不自洽或不可执行措施带入正式 FMEA。",
                reason_tags=tags,
                priority="high" if finding.blocking else "medium",
                blocking=finding.blocking,
            )
        )
    return items


def build_non_expert_question(reason_tags: list[str], label: str) -> tuple[str, list[str], str, str]:
    options = ["多次出现或已知高风险", "偶发或理论可能", "没有已知历史", "不清楚，需要专家确认"]
    if "input_quality" in reason_tags:
        return (
            f"关于{label}，当前输入是否能代表真实产品和使用场景？",
            options,
            "按通用同类模块经验补齐缺口",
            "可能漏掉关键场景、接口或控制措施，导致风险优先级偏低。",
        )
    if "coverage_gap" in reason_tags:
        return (
            f"`{label}` 这个维度是否适用于当前模块，是否存在必须纳入 FMEA 的风险？",
            ["适用且有历史/测试问题", "适用但暂无问题", "不适用", "不清楚，需要专家确认"],
            "暂按需要评审的潜在覆盖缺口处理",
            "可能漏掉某个生命周期、接口或部件维度的失效模式。",
        )
    if "score_uncertainty" in reason_tags:
        return (
            "现有测试、报警、检验或联锁是否足以在客户受影响前发现该问题？",
            ["能稳定发现", "只能部分发现", "基本发现不了", "不清楚，需要测试/质量确认"],
            "O/D 维持 AI 草稿评分",
            "RPN 可能被低估或高估，措施优先级会受到影响。",
        )
    if "broader_analogy" in reason_tags:
        return (
            "这个相似模块案例的机理、接口和使用条件是否真的适用于当前模块？",
            ["高度相似，可以借用", "部分相似，需要改写", "不适用", "不清楚，需要设计专家确认"],
            "作为类比参考保留",
            "可能把其他模块的问题误当成当前模块风险，或使用了错误控制措施。",
        )
    return (
        "这条风险的范围、评分和控制措施是否符合当前产品事实？",
        options,
        "保留为 AI 草稿",
        "可能影响该行是否应进入正式 FMEA 或 Top 风险。",
    )


def build_gap_confirmation_items(
    input_quality: InputQualityDiagnosis,
    coverage_matrix: list[CoverageMatrixItem],
) -> list[ConfirmationItem]:
    items: list[ConfirmationItem] = []
    for missing in input_quality.missing_critical_inputs[:4]:
        tags = ["input_quality", "non_expert_validation"]
        question, options, default, impact = build_non_expert_question(tags, missing)
        items.append(
            ConfirmationItem(
                scope="输入质量诊断",
                row_key=missing,
                why_confirmation_is_needed="输入缺少会影响 FMEA 完整性或评分置信度的关键信息",
                suggested_reviewer_focus="补充事实输入或确认 AI 默认假设是否可接受",
                reference_type="AI assumption",
                source_cases=[],
                plain_language_question=question,
                why_it_matters="输入缺口会影响风险识别、O/D 校准和措施优先级。",
                suggested_options=options,
                default_assumption=default,
                impact_if_wrong=impact,
                reason_tags=tags,
                priority="high" if input_quality.level == "high_risk_missing_context" else "medium",
                blocking=input_quality.level == "high_risk_missing_context",
            )
        )

    for item in coverage_matrix:
        if item.status == "covered":
            continue
        tags = [*item.reason_tags, "non_expert_validation"]
        question, options, default, impact = build_non_expert_question(tags, item.dimension)
        items.append(
            ConfirmationItem(
                scope="覆盖矩阵审查",
                row_key=item.dimension,
                why_confirmation_is_needed=f"覆盖维度为 {item.status}：{item.evidence}",
                suggested_reviewer_focus=item.review_prompt,
                reference_type="AI coverage review",
                source_cases=[],
                plain_language_question=question,
                why_it_matters="覆盖缺口可能代表遗漏的生命周期、接口或部件风险。",
                suggested_options=options,
                default_assumption=default,
                impact_if_wrong=impact,
                reason_tags=tags,
                priority="high" if item.status == "missing" else "medium",
                blocking=item.status == "missing",
            )
        )
    return items


def row_reason_tags(row: DraftRow) -> list[str]:
    tags: list[str] = []
    reason_text = "；".join(row.confirmation_reasons)
    if "O/D" in reason_text or "评分" in reason_text:
        tags.append("score_uncertainty")
    if row.reference_type != "current module":
        tags.append("broader_analogy" if row.reference_type == "broader analogy" else "family_analogy")
    if row.boundary_scopes:
        tags.append("scope_boundary")
    if "模板或知识库" in reason_text:
        tags.append("template_source")
    if not normalize_space(row.current_controls):
        tags.append("missing_controls")
    if tags:
        tags.append("non_expert_validation")
    return tags


def enriched_confirmation_item_for_row(row: DraftRow) -> ConfirmationItem:
    tags = row_reason_tags(row)
    question, options, default, impact = build_non_expert_question(tags, row_key(row))
    priority = "high" if safe_int(row.rpn) is not None and (safe_int(row.rpn) or 0) >= 150 else "medium"
    return ConfirmationItem(
        scope=row.scope,
        row_key=row_key(row),
        why_confirmation_is_needed="；".join(row.confirmation_reasons),
        suggested_reviewer_focus=row.reviewer_focus,
        reference_type=row.reference_type,
        source_cases=row.source_cases,
        review_comment=row.review_comment,
        plain_language_question=question if tags else "",
        why_it_matters="该判断会影响本行是否适用于当前模块，以及 O/D 与措施优先级是否合理。" if tags else "",
        suggested_options=options if tags else [],
        default_assumption=default if tags else "",
        impact_if_wrong=impact if tags else "",
        reason_tags=tags,
        priority=priority,
        blocking="score_uncertainty" in tags or "scope_boundary" in tags,
    )


def row_key(row: DraftRow) -> str:
    head = normalize_space(row.analysis_object or row.function or row.scope or "未命名对象")
    tail = normalize_space(row.failure_mode or "待补失效模式")
    return f"{head} / {tail}"


def short_text(value: str, limit: int = 80) -> str:
    text = normalize_space(value)
    if len(text) <= limit:
        return text
    return text[: limit - 3] + "..."


def first_action_candidate(actions: str) -> str:
    for line in actions.splitlines():
        cleaned = normalize_space(line)
        if cleaned:
            return short_text(cleaned, limit=80)
    return ""


def build_confirmation_queue(
    scope_rows: dict[str, list[DraftRow]],
    input_quality: InputQualityDiagnosis | None = None,
    coverage_matrix: list[CoverageMatrixItem] | None = None,
    quality_gate_findings: list[QualityGateFinding] | None = None,
) -> list[ConfirmationItem]:
    items: list[ConfirmationItem] = []
    if input_quality and coverage_matrix is not None:
        items.extend(build_gap_confirmation_items(input_quality, coverage_matrix))
    if quality_gate_findings:
        items.extend(build_quality_gate_confirmation_items(quality_gate_findings))
    for rows in scope_rows.values():
        for row in rows:
            if row.confirmation_status != "needs expert confirmation":
                continue
            items.append(enriched_confirmation_item_for_row(row))
    return items


def build_top_risks(scope_rows: dict[str, list[DraftRow]]) -> list[dict[str, str]]:
    rows = sorted(
        [row for rows in scope_rows.values() for row in rows if safe_int(row.rpn) is not None],
        key=lambda item: (-(safe_int(item.rpn) or 0), -item.max_match_score),
    )[:8]
    return [
        {
            "scope": row.scope,
            "row_key": row_key(row),
            "failure_mode": row.failure_mode,
            "current_rpn": row.rpn,
            "why_it_matters": short_text(row.effect, 100),
            "first_action_candidate": first_action_candidate(row.recommended_actions),
            "reference_type": row.reference_type,
        }
        for row in rows
    ]


def build_suggested_actions(scope_rows: dict[str, list[DraftRow]]) -> list[dict[str, str]]:
    rows = [
        row
        for rows in scope_rows.values()
        for row in rows
        if normalize_space(row.recommended_actions)
    ]
    rows.sort(
        key=lambda item: (
            -(safe_int(item.rpn) or -1),
            item.reference_type != "current module",
            -item.max_match_score,
            row_key(item),
        )
    )
    return [
        {
            "scope": row.scope,
            "row_key": row_key(row),
            "current_rpn": row.rpn,
            "recommended_action": row.recommended_actions,
            "owner": row.owner,
            "target_date": row.target_date,
            "confirmation_status": row.confirmation_status,
            "review_comment": row.review_comment,
            "reference_type": row.reference_type,
            "source_case": "; ".join(row.source_cases),
        }
        for row in rows
    ]


def build_source_trace(scope_rows: dict[str, list[DraftRow]]) -> list[dict[str, Any]]:
    return [
        {
            "scope": row.scope,
            "row_key": row_key(row),
            "reference_type": row.reference_type,
            "source_cases": row.source_cases,
        }
        for rows in scope_rows.values()
        for row in rows
    ]


def aggregate_rows(scope: ScopeDefinition, scopes: list[ScopeDefinition], matches: list[Match], module: str | None) -> list[DraftRow]:
    cache: dict[Path, Any] = {}
    grouped: dict[tuple[str, str, str, str], dict[str, Any]] = {}
    relevant_rows: list[tuple[Match, dict[str, str], int, int]] = []

    for match in matches:
        row = read_match_row(match, cache)
        hits = scope_hit_count(row, scope.query_terms)
        if hits <= 0:
            continue
        focus_score = scope_focus_score(row, scope.query_terms)
        relevant_rows.append((match, row, module_family_rank(match.sheet, module), focus_score))

    restrict_to_family = sum(1 for _, _, family_rank, _ in relevant_rows if family_rank > 0) >= 4

    for match, row, family_rank, focus_score in relevant_rows:
        if restrict_to_family and family_rank <= 0:
            continue
        if family_rank == 1 and focus_score < 2:
            continue
        if family_rank <= 0 and focus_score < 2:
            continue
        if best_scope_name(row, scopes) != scope.name:
            continue
        key = dedupe_key(row)
        if not any(key):
            continue
        entry = grouped.setdefault(
            key,
            {
                "rows": [],
                "matches": [],
            },
        )
        entry["rows"].append(row)
        entry["matches"].append(match)

    draft_rows: list[DraftRow] = []
    for _, bundle in grouped.items():
        rows = bundle["rows"]
        matches_for_row: list[Match] = bundle["matches"]
        analysis_object = combine_unique([first_value(row, "analysis_object") for row in rows])
        function = combine_unique([first_value(row, "function") for row in rows])
        failure_mode = combine_unique([first_value(row, "failure_mode") for row in rows])
        effect = combine_unique([first_value(row, "effect") for row in rows])
        severity = next((value for value in [first_value(row, "severity") for row in rows] if value), "")
        cause = combine_unique([first_value(row, "cause") for row in rows])
        occurrence = next((value for value in [first_value(row, "occurrence") for row in rows] if value), "")
        current_controls = combine_unique([first_value(row, "current_controls") for row in rows])
        detection = next((value for value in [first_value(row, "detection") for row in rows] if value), "")
        rpn = compute_rpn(
            severity,
            occurrence,
            detection,
            next((value for value in [first_value(row, "rpn") for row in rows] if value), ""),
        )
        recommended_actions = combine_unique([first_value(row, "recommended_actions") for row in rows])
        post_action_severity = next((value for value in [first_value(row, "post_action_severity") for row in rows] if value), "")
        post_action_occurrence = next((value for value in [first_value(row, "post_action_occurrence") for row in rows] if value), "")
        post_action_detection = next((value for value in [first_value(row, "post_action_detection") for row in rows] if value), "")
        post_action_rpn = compute_post_action_rpn(
            post_action_severity,
            post_action_occurrence,
            post_action_detection,
            next((value for value in [first_value(row, "post_action_rpn") for row in rows] if value), ""),
        )
        owner = combine_unique([first_value(row, "owner") for row in rows])
        target_date = combine_unique([first_value(row, "target_date") for row in rows])
        source_cases = []
        for match in matches_for_row:
            label = f"{match.workbook} / {match.sheet} / row {match.excel_row}"
            if label not in source_cases:
                source_cases.append(label)
        reference_type = build_reference_type(matches_for_row, module)
        themes = {match.theme for match in matches_for_row}
        boundary_scopes = []
        for row in rows:
            for scope_name in boundary_scope_names(row, scope.name, scopes):
                if scope_name not in boundary_scopes:
                    boundary_scopes.append(scope_name)
        source_severity = severity
        source_occurrence = occurrence
        source_detection = detection
        confirmation_reasons = build_confirmation_reasons(
            reference_type,
            boundary_scopes,
            themes,
            source_severity,
            source_occurrence,
            source_detection,
        )
        confirmation_status = build_confirmation_status(confirmation_reasons)
        rating_basis = build_rating_basis(
            themes,
            reference_type,
            source_severity,
            source_occurrence,
            source_detection,
        )
        severity, occurrence, detection, rpn, filled_missing_scores = fill_missing_draft_scores(
            severity,
            occurrence,
            detection,
            rpn,
        )
        if filled_missing_scores:
            rating_basis = (
                f"{rating_basis}；源案例缺少 {'/'.join(filled_missing_scores)}，"
                f"已填入保守 AI 草稿 S/O/D={severity}/{occurrence}/{detection}，必须专家确认"
            )
        reviewer_focus = build_reviewer_focus(reference_type, boundary_scopes, source_occurrence, source_detection)

        draft_rows.append(
            DraftRow(
                scope=scope.name,
                analysis_object=analysis_object,
                function=function,
                failure_mode=failure_mode,
                effect=effect,
                severity=severity,
                cause=cause,
                occurrence=occurrence,
                current_controls=current_controls,
                detection=detection,
                rpn=rpn,
                recommended_actions=recommended_actions,
                owner=owner,
                target_date=target_date,
                confirmation_status=confirmation_status,
                rating_basis=rating_basis,
                reference_type=reference_type,
                source_cases=source_cases,
                confirmation_reasons=confirmation_reasons,
                reviewer_focus=reviewer_focus,
                boundary_scopes=boundary_scopes,
                post_action_severity=post_action_severity,
                post_action_occurrence=post_action_occurrence,
                post_action_detection=post_action_detection,
                post_action_rpn=post_action_rpn,
                max_match_score=max(match.score for match in matches_for_row),
                max_scope_hits=max(scope_hit_count(row, scope.query_terms) for row in rows),
            )
        )

    draft_rows.sort(
        key=lambda item: (
            -(safe_int(item.rpn) or -1),
            -item.max_scope_hits,
            -item.max_match_score,
            item.analysis_object,
            item.failure_mode,
        )
    )
    return draft_rows


def lifecycle_profile_by_name(scope_name: str, profiles: list[dict[str, Any]] | None = None) -> dict[str, Any]:
    for profile in profiles or LIFECYCLE_COVERAGE_PROFILES:
        if profile["name"] == scope_name:
            return profile
    return {
        "name": scope_name,
        "keywords": [],
        "target_rows": DEFAULT_SCOPE_MIN_ROWS,
        "function_context": "{function}",
        "effect_context": "{effect}",
        "cause_context": "{cause}",
        "control_context": "{controls}",
        "action_context": "{actions}",
    }


def scope_target_rows(
    scopes: list[ScopeDefinition],
    min_rows: int,
    profiles: list[dict[str, Any]] | None = None,
) -> dict[str, int]:
    if not scopes:
        return {}

    active_profiles = profiles or LIFECYCLE_COVERAGE_PROFILES
    rows_remaining = max(
        min_rows,
        sum(int(lifecycle_profile_by_name(scope.name, active_profiles).get("target_rows", DEFAULT_SCOPE_MIN_ROWS)) for scope in scopes),
    )
    scopes_remaining = len(scopes)
    targets: dict[str, int] = {}

    for scope in scopes:
        profile = lifecycle_profile_by_name(scope.name, active_profiles)
        default_target = int(profile.get("target_rows", DEFAULT_SCOPE_MIN_ROWS))
        target_rows = max(default_target, rows_remaining // max(scopes_remaining, 1))
        targets[scope.name] = target_rows
        rows_remaining -= target_rows
        scopes_remaining -= 1

    return targets


def contextualize_profile_text(profile: dict[str, Any], field_name: str, **values: str) -> str:
    template = profile.get(field_name, "{value}")
    normalized = {key: normalize_space(value) or "待补充" for key, value in values.items()}
    return normalize_space(template.format(**normalized))


def rank_seed_rows(matches: list[Match], module: str | None, scope: ScopeDefinition) -> list[tuple[Match, dict[str, str]]]:
    cache: dict[Path, Any] = {}
    ranked: list[tuple[tuple[int, int, int, int, str], Match, dict[str, str]]] = []
    for match in matches:
        if match.theme not in ALLOWED_THEMES:
            continue
        row = read_match_row(match, cache)
        if not any(dedupe_key(row)):
            continue
        family_rank = module_family_rank(match.sheet, module)
        focus_score = scope_focus_score(row, scope.query_terms)
        rpn = safe_int(first_value(row, "rpn")) or 0
        ranked.append(((-family_rank, -focus_score, -rpn, -match.score, match.excel_row), match, row))
    ranked.sort(key=lambda item: item[0])
    return [(match, row) for _, match, row in ranked]


def draft_row_from_seed(scope: ScopeDefinition, profile: dict[str, Any], match: Match, row: dict[str, str], module: str | None) -> DraftRow:
    analysis_object = first_value(row, "analysis_object") or module or scope.name
    function = first_value(row, "function")
    failure_mode = first_value(row, "failure_mode")
    effect = first_value(row, "effect")
    severity = first_value(row, "severity")
    cause = first_value(row, "cause")
    occurrence = first_value(row, "occurrence")
    controls = first_value(row, "current_controls")
    detection = first_value(row, "detection")
    rpn = compute_rpn(severity, occurrence, detection, first_value(row, "rpn"))
    recommended_actions = first_value(row, "recommended_actions")
    post_action_severity = first_value(row, "post_action_severity")
    post_action_occurrence = first_value(row, "post_action_occurrence")
    post_action_detection = first_value(row, "post_action_detection")
    post_action_rpn = compute_post_action_rpn(
        post_action_severity,
        post_action_occurrence,
        post_action_detection,
        first_value(row, "post_action_rpn"),
    )
    reference_type = build_reference_type([match], module)
    source_cases = [f"{match.workbook} / {match.sheet} / row {match.excel_row}"]
    themes = {match.theme}
    source_severity = severity
    source_occurrence = occurrence
    source_detection = detection
    confirmation_reasons = build_confirmation_reasons(reference_type, [], themes, source_severity, source_occurrence, source_detection)
    confirmation_reasons.append(f"{scope.name} 维度为覆盖扩展生成，需要专家确认该生命周期场景、控制措施和评分是否适用")
    rating_basis = build_rating_basis(themes, reference_type, source_severity, source_occurrence, source_detection)
    severity, occurrence, detection, rpn, filled_missing_scores = fill_missing_draft_scores(
        severity,
        occurrence,
        detection,
        rpn,
    )
    if filled_missing_scores:
        rating_basis = (
            f"{rating_basis}；源案例缺少 {'/'.join(filled_missing_scores)}，"
            f"已填入保守 AI 草稿 S/O/D={severity}/{occurrence}/{detection}，必须专家确认"
        )
    rating_basis = f"{rating_basis}；按 {scope.name} 生命周期维度扩展，参考源案例但不视为已确认结论"

    return DraftRow(
        scope=scope.name,
        analysis_object=analysis_object,
        function=contextualize_profile_text(profile, "function_context", function=function),
        failure_mode=failure_mode,
        effect=contextualize_profile_text(profile, "effect_context", effect=effect),
        severity=severity,
        cause=contextualize_profile_text(profile, "cause_context", cause=cause),
        occurrence=occurrence,
        current_controls=contextualize_profile_text(profile, "control_context", controls=controls),
        detection=detection,
        rpn=rpn,
        recommended_actions=contextualize_profile_text(profile, "action_context", actions=recommended_actions),
        owner=first_value(row, "owner"),
        target_date=first_value(row, "target_date"),
        confirmation_status="needs expert confirmation",
        rating_basis=rating_basis,
        reference_type=reference_type,
        source_cases=source_cases,
        confirmation_reasons=confirmation_reasons,
        reviewer_focus=build_reviewer_focus(reference_type, [], source_occurrence, source_detection),
        post_action_severity=post_action_severity,
        post_action_occurrence=post_action_occurrence,
        post_action_detection=post_action_detection,
        post_action_rpn=post_action_rpn,
        max_match_score=match.score,
        max_scope_hits=scope_hit_count(row, scope.query_terms),
    )


def fallback_lifecycle_row(scope: ScopeDefinition, profile: dict[str, Any], module: str, index: int) -> DraftRow:
    pattern = COVERAGE_GAP_GUIDEWORDS[(index - 1) % len(COVERAGE_GAP_GUIDEWORDS)]
    failure_mode = f"{module}{scope.name}{pattern['failure_mode']}"
    severity = pattern["severity"]
    occurrence = pattern["occurrence"]
    detection = pattern["detection"]
    rpn = compute_rpn(severity, occurrence, detection, "")
    confirmation_reasons = [
        f"{scope.name} 维度缺少足够历史案例，需要补充模块实测、现场和维护数据",
        f"覆盖补缺类别为 {pattern['category']}，需专家确认是否适用于当前对象",
    ]
    return DraftRow(
        scope=scope.name,
        analysis_object=module,
        function=contextualize_profile_text(profile, "function_context", function=f"{module}在{scope.name}阶段保持功能和安全边界"),
        failure_mode=failure_mode,
        effect=contextualize_profile_text(profile, "effect_context", effect=pattern["effect"]),
        severity=severity,
        cause=contextualize_profile_text(profile, "cause_context", cause=pattern["cause"]),
        occurrence=occurrence,
        current_controls=contextualize_profile_text(profile, "control_context", controls=pattern["controls"]),
        detection=detection,
        rpn=rpn,
        recommended_actions=contextualize_profile_text(profile, "action_context", actions=pattern["action"]),
        owner="责任工程师待定",
        target_date="待定",
        confirmation_status="needs expert confirmation",
        rating_basis=f"第 {index} 条覆盖补缺行；guideword={pattern['category']}；无足够源案例，已填入保守 AI 草稿 S/O/D={severity}/{occurrence}/{detection}，需人工确认",
        reference_type="broader analogy",
        source_cases=[],
        confirmation_reasons=confirmation_reasons,
        reviewer_focus=f"确认是否需要保留该 {pattern['category']} 风险，并补齐机理、现行控制、S/O/D 与责任人",
    )


def pad_scope_rows_to_minimum(
    module: str,
    scopes: list[ScopeDefinition],
    scope_rows: dict[str, list[DraftRow]],
    min_rows: int,
    profiles: list[dict[str, Any]] | None = None,
) -> dict[str, list[DraftRow]]:
    targets = scope_target_rows(scopes, min_rows, profiles)
    active_profiles = profiles or LIFECYCLE_COVERAGE_PROFILES

    for scope in scopes:
        rows = scope_rows.setdefault(scope.name, [])
        profile = lifecycle_profile_by_name(scope.name, active_profiles)
        while len(rows) < targets.get(scope.name, DEFAULT_SCOPE_MIN_ROWS):
            rows.append(fallback_lifecycle_row(scope, profile, module, len(rows) + 1))
        rows.sort(
            key=lambda item: (
                -(safe_int(item.rpn) or -1),
                -item.max_match_score,
                item.analysis_object,
                item.failure_mode,
            )
        )

    return scope_rows


def build_lifecycle_coverage_rows(
    module: str,
    input_text: str,
    extracted_terms: list[str],
    scopes: list[ScopeDefinition],
    min_rows: int,
    profiles: list[dict[str, Any]] | None = None,
) -> dict[str, list[DraftRow]]:
    scope_rows: dict[str, list[DraftRow]] = {}
    base_query_terms = [module, *extracted_terms[:12], *tokenize(input_text)[:24]]
    active_profiles = profiles or LIFECYCLE_COVERAGE_PROFILES
    targets = scope_target_rows(scopes, min_rows, active_profiles)

    for scope in scopes:
        profile = lifecycle_profile_by_name(scope.name, active_profiles)
        target_rows = targets.get(scope.name, DEFAULT_SCOPE_MIN_ROWS)

        query = " ".join([*base_query_terms, *scope.query_terms, *profile.get("keywords", [])])
        seed_rows = rank_seed_rows(collect_matches(query, module), module, scope)

        used_keys: set[tuple[str, str]] = set()
        draft_rows: list[DraftRow] = []
        for match, row in seed_rows:
            key = (first_value(row, "analysis_object"), first_value(row, "failure_mode"))
            if key in used_keys:
                continue
            used_keys.add(key)
            draft_rows.append(draft_row_from_seed(scope, profile, match, row, module))
            if len(draft_rows) >= target_rows:
                break

        while len(draft_rows) < target_rows:
            draft_rows.append(fallback_lifecycle_row(scope, profile, module, len(draft_rows) + 1))

        draft_rows.sort(
            key=lambda item: (
                -(safe_int(item.rpn) or -1),
                -item.max_match_score,
                item.analysis_object,
                item.failure_mode,
            )
        )
        scope_rows[scope.name] = draft_rows

    return scope_rows


def format_md_cell(value: str) -> str:
    return value.replace("|", "\\|").replace("\n", "<br>")


def capture_row_style(ws: Any, source_row: int, min_col: int = 2, max_col: int = 23) -> dict[str, Any]:
    row_dimension = ws.row_dimensions[source_row]
    return {
        "height": row_dimension.height,
        "hidden": row_dimension.hidden,
        "outlineLevel": row_dimension.outlineLevel,
        "cells": [
            {
                "_style": copy(ws.cell(row=source_row, column=column)._style),
                "font": copy(ws.cell(row=source_row, column=column).font),
                "fill": copy(ws.cell(row=source_row, column=column).fill),
                "border": copy(ws.cell(row=source_row, column=column).border),
                "alignment": copy(ws.cell(row=source_row, column=column).alignment),
                "number_format": ws.cell(row=source_row, column=column).number_format,
                "protection": copy(ws.cell(row=source_row, column=column).protection),
            }
            for column in range(min_col, max_col + 1)
        ],
    }


def apply_row_style_snapshot(ws: Any, snapshot: dict[str, Any], target_row: int, min_col: int = 2) -> None:
    target_dimension = ws.row_dimensions[target_row]
    target_dimension.height = snapshot["height"]
    target_dimension.hidden = snapshot["hidden"]
    target_dimension.outlineLevel = snapshot["outlineLevel"]

    for column_offset, style in enumerate(snapshot["cells"], start=min_col):
        target = ws.cell(row=target_row, column=column_offset)
        target._style = copy(style["_style"])
        target.font = copy(style["font"])
        target.fill = copy(style["fill"])
        target.border = copy(style["border"])
        target.alignment = copy(style["alignment"])
        target.number_format = style["number_format"]
        target.protection = copy(style["protection"])


def set_if_sheet_cell(workbook: Any, sheet_name: str, cell_ref: str, value: Any) -> None:
    if sheet_name in workbook.sheetnames:
        workbook[sheet_name][cell_ref] = value


def summarize_input_for_cover(input_text: str, limit: int = 180) -> str:
    summary = normalize_space(input_text)
    if len(summary) <= limit:
        return summary
    return summary[:limit].rstrip() + "..."


def build_rating_basis_cell(row: DraftRow) -> str:
    parts = [row.rating_basis.strip()]
    if row.confirmation_status:
        parts.append(f"确认状态：{row.confirmation_status}")
    if row.reference_type:
        parts.append(f"参考类型：{row.reference_type}")
    if row.review_comment:
        parts.append(f"评审备注：{row.review_comment}")
    if row.source_cases:
        parts.append(f"来源：{'; '.join(row.source_cases[:3])}")
    return "\n".join(part for part in parts if part)


def extract_parameter_indicators(*texts: str, limit: int = 3) -> str:
    candidates: list[str] = []
    pattern = re.compile(
        r"[^。；;\n，,]*("
        r"(?:≤|≥|<|>|±|=|不超过|不少于|小于|大于)"
        r"\s*[\d.]+|[\d.]+\s*(?:ms|μs|us|s|Hz|kHz|MHz|GHz|W|kW|V|mV|A|mA|℃|°C|%|Ω|ohm|dB|ppm|nm|μm|um|Pa)"
        r")[^。；;\n，,]*"
    )
    for text in texts:
        for match in pattern.finditer(text or ""):
            candidate = normalize_space(match.group(0))
            if candidate and candidate not in candidates:
                candidates.append(candidate)
            if len(candidates) >= limit:
                return "；".join(candidates)
    return "；".join(candidates)


def split_current_controls(controls: str) -> tuple[str, str]:
    text = normalize_space(controls)
    if not text:
        return "", ""

    prevention = ""
    detection = ""
    prevention_match = re.search(r"预防[:：]\s*(.*?)(?:探测[:：]|$)", text)
    detection_match = re.search(r"探测[:：]\s*(.*)$", text)
    if prevention_match:
        prevention = normalize_space(prevention_match.group(1))
    if detection_match:
        detection = normalize_space(detection_match.group(1))
    if prevention or detection:
        return prevention or text, detection or text

    return text, text


def infer_owner(row: DraftRow) -> str:
    text = " ".join([row.scope, row.analysis_object, row.function, row.failure_mode, row.cause])
    owner_rules = [
        ("软件", "软件工程师"),
        ("算法", "算法工程师"),
        ("逻辑", "控制逻辑工程师"),
        ("热", "热设计工程师"),
        ("温度", "热设计工程师"),
        ("散热", "热设计工程师"),
        ("结构", "结构工程师"),
        ("机械", "机械工程师"),
        ("接地", "电气工程师"),
        ("电源", "电气工程师"),
        ("EMC", "EMC工程师"),
        ("电磁", "EMC工程师"),
        ("射频", "射频工程师"),
        ("功率", "射频工程师"),
        ("保护", "系统工程师"),
        ("测试", "测试工程师"),
        ("物流", "供应链/物流工程师"),
        ("运输", "供应链/物流工程师"),
        ("维护", "客户服务工程师"),
        ("保养", "客户服务工程师"),
    ]
    for keyword, owner in owner_rules:
        if keyword in text:
            return owner
    return "责任工程师待定"


def build_template_row_values(index: int, module: str, row: DraftRow) -> list[Any]:
    prevention_controls, detection_controls = split_current_controls(row.current_controls)
    parameter_indicators = extract_parameter_indicators(row.function, row.effect, row.cause, row.recommended_actions)
    source_traces = "; ".join(row.source_cases) if row.source_cases else ""
    return [
        index,                              # col 2: 序号
        row.scope,                          # col 3: Scope path
        row.analysis_object or module,      # col 4: Leaf 节点
        row.function,                       # col 5: Analysis object
        parameter_indicators,               # col 6: Function or requirement
        row.effect,                         # col 7: P-Diagram 锚点
        row.severity,                       # col 8: Failure mode
        row.failure_mode,                   # col 9: Failure mode canonical
        row.cause,                          # col 10: Failure effect
        prevention_controls,                # col 11: S
        row.occurrence,                     # col 12: Cause or mechanism
        detection_controls,                 # col 13: O
        row.detection,                      # col 14: Current controls (prevention)
        row.rpn,                            # col 15: Current controls (detection) — overridden with formula
        build_rating_basis_cell(row),       # col 16: D
        row.recommended_actions,            # col 17: RPN
        row.owner or infer_owner(row),      # col 18: Recommended actions
        row.target_date or "待定",           # col 19: Owner
        row.post_action_severity,           # col 20: Target date
        row.post_action_occurrence,         # col 21: 改进后 S
        row.post_action_detection,          # col 22: 改进后 O
        row.post_action_rpn,                # col 23: 改进后 D
        None,                               # col 24: 改进后 RPN — overridden with formula T*U*V
        # M2 new-column defaults: legacy import / AI-draft has no evidence grading info
        "ai-inferred",                      # col 25: Evidence grade
        None,                               # col 26: Confidence
        "",                                 # col 27: Confidence breakdown
        "N",                                # col 28: Multi-role corroborated
        "",                                 # col 29: Rating history
        "Y",                                # col 30: Needs human confirmation
        source_traces,                      # col 31: Source traces
        "",                                 # col 32: AI 打分推导依据
    ]


def template_fmea_rows(
    module: str,
    scopes: list[ScopeDefinition],
    scope_rows: dict[str, list[DraftRow]],
) -> list[list[Any]]:
    output_rows: list[list[Any]] = []
    index = 1
    for scope in scopes:
        rows = scope_rows.get(scope.name, [])
        if not rows:
            output_rows.append(
                [
                    index,                  # col 2: 序号
                    scope.name,             # col 3: Scope path
                    module,                 # col 4: Leaf 节点
                    "",                     # col 5: Analysis object
                    "",                     # col 6: Function or requirement
                    "",                     # col 7: P-Diagram 锚点
                    "",                     # col 8: Failure mode
                    "",                     # col 9: Failure mode canonical
                    "",                     # col 10: Failure effect
                    "",                     # col 11: S
                    "",                     # col 12: Cause or mechanism
                    "",                     # col 13: O
                    "",                     # col 14: Current controls (prevention)
                    "",                     # col 15: Current controls (detection)
                    "未召回到足够案例，建议补充输入或手工定义 scope。",  # col 16: D
                    "补充模块功能、失效模式、S/O/D 评分依据和现行控制。",  # col 17: RPN
                    "",                     # col 18: Recommended actions
                    "",                     # col 19: Owner
                    "",                     # col 20: Target date
                    "",                     # col 21: 改进后 S
                    "",                     # col 22: 改进后 O
                    "",                     # col 23: 改进后 D
                    None,                   # col 24: 改进后 RPN — overridden with formula T*U*V
                    # M2 new-column defaults
                    "ai-inferred",          # col 25: Evidence grade
                    None,                   # col 26: Confidence
                    "",                     # col 27: Confidence breakdown
                    "N",                    # col 28: Multi-role corroborated
                    "",                     # col 29: Rating history
                    "Y",                    # col 30: Needs human confirmation
                    "",                     # col 31: Source traces
                    "",                     # col 32: AI 打分推导依据
                ]
            )
            index += 1
            continue
        for row in rows:
            output_rows.append(build_template_row_values(index, module, row))
            index += 1
    return output_rows


def render_template_cover(
    workbook: Any,
    module: str,
    fmea_type: str,
    input_text: str,
    scopes: list[ScopeDefinition],
    scope_rows: dict[str, list[DraftRow]],
) -> None:
    if "封面" not in workbook.sheetnames:
        return

    sheet_count_text = " / ".join(scope.name for scope in scopes[:8])
    if len(scopes) > 8:
        sheet_count_text += f" / 等 {len(scopes)} 个范围"
    indicators = extract_parameter_indicators(input_text, limit=4)
    metadata = FMEA_TYPE_METADATA[normalize_fmea_type(fmea_type)]
    if not indicators:
        indicators = metadata["indicator_fallback"].format(module=module or "当前对象")

    set_if_sheet_cell(workbook, "封面", "B2", f"{module or '未命名模块'} {fmea_type}分析报告")
    set_if_sheet_cell(workbook, "封面", "B3", metadata["subtitle"].format(module=module or "Current Module", module_en=module or "Current Module"))
    set_if_sheet_cell(workbook, "封面", "C6", module or "未指定")
    set_if_sheet_cell(workbook, "封面", "C7", indicators)
    set_if_sheet_cell(workbook, "封面", "C8", metadata["standard_note"])
    set_if_sheet_cell(workbook, "封面", "C9", sheet_count_text or "未拆分")
    set_if_sheet_cell(workbook, "封面", "C10", "历史FMEA案例库 / 相邻模块类比 / 当前输入约束")
    set_if_sheet_cell(workbook, "封面", "C11", date.today().isoformat())
    set_if_sheet_cell(workbook, "封面", "C12", "V1.0")

    if workbook["封面"]["C17"].value:
        workbook["封面"]["C17"] = f"按 template.xlsx 样式输出的 {sum(len(rows) for rows in scope_rows.values())} 条 FMEA 草稿记录"


def render_template_fmea_sheet(workbook: Any, module: str, scopes: list[ScopeDefinition], scope_rows: dict[str, list[DraftRow]]) -> None:
    if "FMEA主表" not in workbook.sheetnames:
        raise ValueError("标准输出模板中缺少必需工作表：FMEA主表")

    ws = workbook["FMEA主表"]
    if ws["B2"].value != "序号":
        raise ValueError("标准输出模板的 FMEA主表 必须保持标准格式：B2:W2 为表头，B列为序号")

    template_odd_row = 3 if ws.max_row >= 3 else 2
    template_even_row = 4 if ws.max_row >= 4 else template_odd_row
    data_start_row = 3
    min_col = 2
    max_col = 32
    odd_style = capture_row_style(ws, template_odd_row, min_col=min_col, max_col=max_col)
    even_style = capture_row_style(ws, template_even_row, min_col=min_col, max_col=max_col)

    if ws.max_row >= data_start_row:
        ws.delete_rows(data_start_row, ws.max_row - data_start_row + 1)

    for row_offset, values in enumerate(template_fmea_rows(module, scopes, scope_rows), start=0):
        target_row = data_start_row + row_offset
        row_style = odd_style if row_offset % 2 == 0 else even_style
        apply_row_style_snapshot(ws, row_style, target_row, min_col=min_col)
        for column_offset, value in enumerate(values, start=min_col):
            ws.cell(row=target_row, column=column_offset, value=value)
        ws.cell(row=target_row, column=15, value=f"=H{target_row}*L{target_row}*N{target_row}")
        ws.cell(row=target_row, column=24, value=f"=T{target_row}*U{target_row}*V{target_row}")

    if ws.max_row >= 2:
        ws.auto_filter.ref = f"B2:AF{ws.max_row}"
    apply_quality_gate_excel_format(ws, ws.max_row)


def apply_quality_gate_excel_format(ws: Any, last_row: int) -> None:
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    header_fill = PatternFill("solid", fgColor="333333")
    header_font = Font(color="FFFFFF", bold=True)
    serial_fill = PatternFill("solid", fgColor="D9E1F2")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    text_align = Alignment(horizontal="left", vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    numeric_columns = {8, 12, 14, 15, 20, 21, 22, 24, 26}

    for col_idx in range(2, 33):
        cell = ws.cell(row=2, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

    for row_idx in range(3, max(last_row, 3) + 1):
        for col_idx in range(2, 33):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            if col_idx == 2:
                cell.fill = serial_fill
                cell.font = Font(bold=True)
                cell.alignment = center_align
            elif col_idx in numeric_columns:
                cell.alignment = center_align
            else:
                cell.alignment = text_align


def render_excel_workbook(
    module: str,
    fmea_type: str,
    input_text: str,
    scopes: list[ScopeDefinition],
    scope_rows: dict[str, list[DraftRow]],
    excel_path: Path,
) -> None:
    template_path = template_path_for(fmea_type)
    if not template_path.exists():
        raise FileNotFoundError(f"未找到 Excel 输出模板：{template_path}")

    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as exc:
        raise ModuleNotFoundError("Excel output requires the optional dependency `openpyxl`.") from exc

    workbook = load_workbook(template_path)
    render_template_cover(workbook, module, fmea_type, input_text, scopes, scope_rows)
    render_template_fmea_sheet(workbook, module, scopes, scope_rows)

    excel_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(excel_path)


def render_markdown(
    module: str,
    fmea_type: str,
    input_text: str,
    scopes: list[ScopeDefinition],
    scope_rows: dict[str, list[DraftRow]],
) -> str:
    input_quality = diagnose_input_quality(module, fmea_type, input_text, scopes)
    coverage_matrix = build_coverage_matrix(fmea_type, scopes, scope_rows, input_quality)
    quality_gate_findings = build_quality_gate_findings(fmea_type, scope_rows)
    confirmation_queue = build_confirmation_queue(scope_rows, input_quality, coverage_matrix, quality_gate_findings)
    top_risks = build_top_risks(scope_rows)
    suggested_actions = build_suggested_actions(scope_rows)
    source_trace = build_source_trace(scope_rows)
    lines = [
        f"# {module or '未命名模块'} 首版 {fmea_type} 草稿",
        "",
        "- 生成方式: `draft_fmea_from_cases.py`",
        f"- 模块: `{module}`" if module else "- 模块: 未指定",
        f"- FMEA 类型: `{fmea_type}`",
        f"- 输入长度: `{len(input_text)}` 字符",
        "",
        "## 输入摘要",
        "",
        "> " + format_md_cell(input_text[:400] + ("..." if len(input_text) > 400 else "")),
        "",
        "## Input Quality Diagnosis",
        "",
        f"- Level: `{input_quality.level}`",
        f"- Summary: {input_quality.summary}",
        "",
        "| Signal | Status | Evidence | Missing detail |",
        "| --- | --- | --- | --- |",
    ]
    for signal in input_quality.signals:
        lines.append(
            f"| {format_md_cell(signal.signal)} | {format_md_cell(signal.status)} | {format_md_cell(signal.evidence)} | {format_md_cell(signal.missing_detail)} |"
        )
    lines.extend(
        [
            "",
            "## Coverage Matrix Review",
            "",
            "| Dimension | Status | Evidence | Review prompt |",
            "| --- | --- | --- | --- |",
        ]
    )
    for item in coverage_matrix:
        lines.append(
            f"| {format_md_cell(item.dimension)} | {format_md_cell(item.status)} | {format_md_cell(item.evidence)} | {format_md_cell(item.review_prompt)} |"
        )
    lines.extend(
        [
            "",
            "## Quality Gate Findings",
            "",
            "| Gate | Status | Row key | Finding | Required fix or confirmation | Blocking | Reason tags |",
            "| --- | --- | --- | --- | --- | --- | --- |",
        ]
    )
    if not quality_gate_findings:
        lines.append("|  | pass |  | 未发现额外质量门禁问题,仍需专家评审。 |  | False |  |")
    else:
        for item in quality_gate_findings[:24]:
            lines.append(
                f"| {format_md_cell(item.gate)} | {format_md_cell(item.status)} | {format_md_cell(item.row_key)} | {format_md_cell(item.finding)} | {format_md_cell(item.required_fix_or_confirmation)} | {format_md_cell(str(item.blocking))} | {format_md_cell(' / '.join(item.reason_tags))} |"
            )
        if len(quality_gate_findings) > 24:
            lines.append(
                f"|  | review |  | 还有 {len(quality_gate_findings) - 24} 条质量门禁发现未在 Markdown 预览中展开。 | 查看 JSON companion 的 `quality_gate_findings`。 | False | quality_gate |"
            )
    lines.extend(
        [
            "",
            "## Scope 规划",
            "",
            "| Scope | 检索关键词 | 来源 | 命中数 | 说明 |",
            "| --- | --- | --- | ---: | --- |",
        ]
    )
    for scope in scopes:
        terms = scope.query_terms or scope.extracted_terms
        source = "auto" if scope.auto_suggested else "manual"
        lines.append(
            f"| {format_md_cell(scope.name)} | {format_md_cell(' / '.join(terms))} | {source} | {scope.hit_count} | {format_md_cell(scope.reason)} |"
        )

    for scope in scopes:
        rows = scope_rows.get(scope.name, [])
        lines.extend(
            [
                "",
                f"## {scope.name}",
                "",
                "| Scope | Analysis object | Function or requirement | Failure mode | Failure effect | S | Cause or mechanism | O | Current controls | D | RPN | Recommended actions | Confirmation status | Review comment | Rating basis | Reference type | Source case |",
                "| --- | --- | --- | --- | --- | ---: | --- | ---: | --- | ---: | ---: | --- | --- | --- | --- | --- | --- |",
            ]
        )
        if not rows:
            lines.append("|  |  |  |  |  |  |  |  |  |  |  | 未召回到足够案例，建议补充输入或手工定义 scope。 | needs expert confirmation |  | 需要补充输入后再判断 | broader analogy |  |")
            continue
        for row in rows:
            lines.append(
                "| "
                + " | ".join(
                    [
                        format_md_cell(row.scope),
                        format_md_cell(row.analysis_object),
                        format_md_cell(row.function),
                        format_md_cell(row.failure_mode),
                        format_md_cell(row.effect),
                        format_md_cell(row.severity),
                        format_md_cell(row.cause),
                        format_md_cell(row.occurrence),
                        format_md_cell(row.current_controls),
                        format_md_cell(row.detection),
                        format_md_cell(row.rpn),
                        format_md_cell(row.recommended_actions),
                        format_md_cell(row.confirmation_status),
                        format_md_cell(row.review_comment),
                        format_md_cell(row.rating_basis),
                        format_md_cell(row.reference_type),
                        format_md_cell("; ".join(row.source_cases)),
                    ]
                )
                + " |"
            )

    lines.extend(
        [
            "",
            "## Top Risks",
            "",
            "| Scope | Row key | Failure mode | Current RPN | Why it matters | First action candidate | Reference type |",
            "| --- | --- | --- | ---: | --- | --- | --- |",
        ]
    )
    for item in top_risks:
        lines.append(
            f"| {format_md_cell(item['scope'])} | {format_md_cell(item['row_key'])} | {format_md_cell(item['failure_mode'])} | {format_md_cell(item['current_rpn'])} | {format_md_cell(item['why_it_matters'])} | {format_md_cell(item['first_action_candidate'])} | {format_md_cell(item['reference_type'])} |"
        )

    lines.extend(
        [
            "",
            "## Rows Needing Confirmation",
            "",
            "| Scope | Row key | Why confirmation is needed | Plain-language question | Suggested options | Impact if wrong | Priority | Blocking | Suggested reviewer focus | Review comment | Reference type | Source case |",
            "| --- | --- | --- | --- | --- | --- | --- | --- | --- | --- | --- | --- |",
        ]
    )
    if not confirmation_queue:
        lines.append("|  |  | 当前没有额外确认队列，仍建议在评审中校准 O/D。 |  |  |  |  |  |  |  |  |  |")
    else:
        for item in confirmation_queue[:12]:
            lines.append(
                f"| {format_md_cell(item.scope)} | {format_md_cell(item.row_key)} | {format_md_cell(item.why_confirmation_is_needed)} | {format_md_cell(item.plain_language_question)} | {format_md_cell(' / '.join(item.suggested_options))} | {format_md_cell(item.impact_if_wrong)} | {format_md_cell(item.priority)} | {format_md_cell(str(item.blocking))} | {format_md_cell(item.suggested_reviewer_focus)} | {format_md_cell(item.review_comment)} | {format_md_cell(item.reference_type)} | {format_md_cell('; '.join(item.source_cases))} |"
            )

    lines.extend(
        [
            "",
            "## Suggested Actions",
            "",
            "| Scope | Row key | Current RPN | Recommended action | Owner | Target date | Confirmation status | Review comment | Reference type | Source case |",
            "| --- | --- | ---: | --- | --- | --- | --- | --- | --- | --- |",
        ]
    )
    for item in suggested_actions[:12]:
        lines.append(
            f"| {format_md_cell(item['scope'])} | {format_md_cell(item['row_key'])} | {format_md_cell(item['current_rpn'])} | {format_md_cell(item['recommended_action'])} | {format_md_cell(item['owner'])} | {format_md_cell(item['target_date'])} | {format_md_cell(item['confirmation_status'])} | {format_md_cell(item['review_comment'])} | {format_md_cell(item['reference_type'])} | {format_md_cell(item['source_case'])} |"
        )

    lines.extend(
        [
            "",
            "## Source Trace",
            "",
            "| Scope | Row key | Reference type | Source case |",
            "| --- | --- | --- | --- |",
        ]
    )
    for item in source_trace:
        lines.append(
            f"| {format_md_cell(item['scope'])} | {format_md_cell(item['row_key'])} | {format_md_cell(item['reference_type'])} | {format_md_cell('; '.join(item['source_cases']))} |"
        )

    return "\n".join(lines) + "\n"


def build_json_payload(
    module: str,
    fmea_type: str,
    input_text: str,
    scopes: list[ScopeDefinition],
    scope_rows: dict[str, list[DraftRow]],
) -> dict[str, Any]:
    input_quality = diagnose_input_quality(module, fmea_type, input_text, scopes)
    coverage_matrix = build_coverage_matrix(fmea_type, scopes, scope_rows, input_quality)
    quality_gate_findings = build_quality_gate_findings(fmea_type, scope_rows)
    confirmation_queue = build_confirmation_queue(scope_rows, input_quality, coverage_matrix, quality_gate_findings)
    top_risks = build_top_risks(scope_rows)
    suggested_actions = build_suggested_actions(scope_rows)
    source_trace = build_source_trace(scope_rows)
    return {
        "module": module,
        "fmea_type": fmea_type,
        "input_text": input_text,
        "input_quality_diagnosis": asdict(input_quality),
        "coverage_matrix": [asdict(item) for item in coverage_matrix],
        "quality_gate_findings": [asdict(item) for item in quality_gate_findings],
        "scopes": [asdict(scope) for scope in scopes],
        "rows": [asdict(row) for rows in scope_rows.values() for row in rows],
        "confirmation_queue": [asdict(item) for item in confirmation_queue],
        "top_risks": top_risks,
        "suggested_actions": suggested_actions,
        "source_trace": source_trace,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Draft a first FMEA table from natural-language input and historical cases.")
    parser.add_argument("--module", required=True, help="Module name or analysis object.")
    parser.add_argument("--fmea-type", default="DFMEA", choices=sorted(VALID_FMEA_TYPES), help="FMEA type, default is DFMEA.")
    parser.add_argument("--input-file", help="Path to a UTF-8 text file containing the natural-language input.")
    parser.add_argument("--input-text", help="Natural-language input text.")
    parser.add_argument(
        "--scope",
        action="append",
        default=[],
        help="Optional scope definition in the form 'Scope Name::keyword1 keyword2'. Repeatable.",
    )
    parser.add_argument("--top-k", type=int, default=30, help="Number of source matches to keep per scope.")
    parser.add_argument(
        "--coverage-mode",
        choices=["lifecycle", "subsystem", "part"],
        default="lifecycle",
        help="Use type-specific rich coverage by default; choose subsystem for narrow grouping or part for DFMEA part-level coverage.",
    )
    parser.add_argument(
        "--min-rows",
        type=int,
        default=None,
        help="Minimum FMEA rows to draft in rich coverage mode; defaults are type-specific.",
    )
    parser.add_argument("--excel-out", help="Optional path to save the generated Excel workbook.")
    parser.add_argument("--markdown-out", help="Optional path to save the generated Markdown draft.")
    parser.add_argument("--json-out", help="Optional path to save the generated JSON draft.")
    args = parser.parse_args()

    fmea_type = normalize_fmea_type(args.fmea_type)
    coverage_profiles = coverage_profiles_for(fmea_type, args.coverage_mode)
    min_rows = args.min_rows if args.min_rows is not None else int(FMEA_TYPE_METADATA[fmea_type]["default_min_rows"])
    input_text = load_input_text(args)
    extracted_terms = extract_query_terms(input_text, args.module)

    scopes = [parse_scope(raw_scope) for raw_scope in args.scope]
    use_lifecycle_coverage = args.coverage_mode in {"lifecycle", "part"} and not scopes
    if use_lifecycle_coverage:
        scopes = suggest_lifecycle_scopes(args.module, input_text, extracted_terms, coverage_profiles)
    elif not scopes:
        scopes = suggest_scopes(args.module, input_text, extracted_terms)
        if not scopes:
            scopes = [
                ScopeDefinition(
                    name=f"{args.module}整体范围",
                    query_terms=extracted_terms[:12],
                    extracted_terms=extracted_terms,
                    auto_suggested=True,
                    hit_count=0,
                    reason="fallback: no strong scope profile matched",
                )
            ]
    else:
        for scope in scopes:
            scope.extracted_terms = [term for term in extracted_terms if term in scope.query_terms or term in input_text]

    if use_lifecycle_coverage:
        scope_rows = build_lifecycle_coverage_rows(args.module, input_text, extracted_terms, scopes, min_rows, coverage_profiles)
    else:
        scope_rows: dict[str, list[DraftRow]] = {}
        for scope in scopes:
            query = " ".join(scope.query_terms or scope.extracted_terms)
            matches = collect_matches(query, args.module)
            matches = [match for match in matches if match.theme in ALLOWED_THEMES][: args.top_k]
            scope_rows[scope.name] = aggregate_rows(scope, scopes, matches, args.module)
        scope_rows = pad_scope_rows_to_minimum(args.module, scopes, scope_rows, min_rows, coverage_profiles)

    markdown = render_markdown(args.module, fmea_type, input_text, scopes, scope_rows)
    payload = build_json_payload(args.module, fmea_type, input_text, scopes, scope_rows)

    if args.excel_out:
        excel_path = Path(args.excel_out)
        render_excel_workbook(args.module, fmea_type, input_text, scopes, scope_rows, excel_path)

    if args.markdown_out:
        markdown_path = Path(args.markdown_out)
        markdown_path.parent.mkdir(parents=True, exist_ok=True)
        markdown_path.write_text(markdown, encoding="utf-8")
    else:
        print(markdown)

    if args.json_out:
        json_path = Path(args.json_out)
        json_path.parent.mkdir(parents=True, exist_ok=True)
        json_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


if __name__ == "__main__":
    main()
