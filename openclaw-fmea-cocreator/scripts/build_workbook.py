"""Render fmea_normalized.json into the 31-column 5-sheet template.xlsx."""
from __future__ import annotations

import argparse
import json
import shutil
from datetime import date
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule, FormulaRule
from openpyxl.utils import get_column_letter

SKILL_DIR = Path(__file__).resolve().parent.parent
TEMPLATE = SKILL_DIR / "template.xlsx"

EVIDENCE_COLOR = {
    "evidence-backed": "C6EFCE",
    "historical-supported": "E2EFDA",
    "multi-role-inferred": "FFF2CC",
    "ai-inferred": "FFD966",
    "contradicted": "F4B084",
}

HEADER_FILL = PatternFill("solid", fgColor="333333")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SERIAL_FILL = PatternFill("solid", fgColor="D9E1F2")
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
TEXT_ALIGN = Alignment(horizontal="left", vertical="top", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
NUMERIC_COLUMNS = {11, 13, 16, 17, 21, 22, 23, 24, 26}


def _walk_leaves(node):
    if not node:
        return
    if node.get("level") == "component":
        yield node
    for child in node.get("children", []):
        yield from _walk_leaves(child)


def _render_cover(ws, normalized: dict, structure: dict | None) -> None:
    ws["B2"] = f"{normalized['module_root']} {normalized['fmea_type']}分析报告"
    ws["C6"] = normalized["module_root"]
    ws["C9"] = ", ".join(p["scope_id"] for p in (structure or {}).get("p_diagrams", [])) or "未拆分"
    ws["C10"] = "历史FMEA案例库 + 多角色 LLM 推理"
    ws["C11"] = date.today().isoformat()
    ws["C12"] = "V0.3.0-m2"
    rows = normalized.get("rows", [])
    ws["C15"] = sum(1 for _ in _walk_leaves((structure or {}).get("hierarchy", {}))) if structure else len(rows)
    ws["C16"] = len(normalized.get("coverage_gaps", []))
    grade_counts = {}
    for r in rows:
        grade_counts[r["evidence_grade"]] = grade_counts.get(r["evidence_grade"], 0) + 1
    ws["C17"] = " | ".join(f"{k}={v}" for k, v in grade_counts.items())
    if rows:
        avg_conf = sum(r["confidence"] for r in rows) / len(rows)
        ws["C18"] = f"平均置信度 {avg_conf:.2f}, 行数 {len(rows)}"


def _render_main(ws, rows: list) -> None:
    for idx, row in enumerate(rows, start=1):
        excel_row = idx + 2
        values = [
            idx,
            row["scope_path"],
            row["leaf_id"],
            row.get("leaf_id"),
            "",
            row["p_diagram_anchor"],
            row["failure_mode"],
            row["failure_mode_canonical"],
            f"客户:{row['effect_customer']} | 系统:{row['effect_system']}",
            row["severity"],
            row["cause"],
            row["occurrence"],
            row["current_controls_prevention"],
            row["current_controls_detection"],
            row["detection"],
            row["rpn"],
            "; ".join(row["recommended_actions"]),
            "",
            "",
            "", "", "", "",
            row["evidence_grade"],
            row["confidence"],
            json.dumps(row["confidence_breakdown"], ensure_ascii=False),
            "Y" if row.get("multi_role_corroborated") else "N",
            json.dumps(row.get("rating_history", {}), ensure_ascii=False),
            "Y" if row.get("needs_human_confirmation") else "N",
            "; ".join(t.get("ref") or t.get("role", "") for t in row.get("source_traces", [])),
            "",
        ]
        for col_idx, value in enumerate(values, start=2):
            ws.cell(row=excel_row, column=col_idx, value=value)
    last_row = len(rows) + 2
    if last_row >= 3:
        for grade, color in EVIDENCE_COLOR.items():
            rule = FormulaRule(formula=[f'$Y3="{grade}"'], stopIfTrue=False, fill=PatternFill("solid", fgColor=color))
            ws.conditional_formatting.add(f"Y3:Y{last_row}", rule)
        ws.conditional_formatting.add(
            f"Z3:Z{last_row}",
            DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color="638EC6", showValue=True),
        )
    _apply_fmea_table_format(ws, last_row)


def _apply_fmea_table_format(ws, last_row: int) -> None:
    for col_idx in range(2, 33):
        cell = ws.cell(row=2, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER

    for row_idx in range(3, max(last_row, 3) + 1):
        for col_idx in range(2, 33):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = THIN_BORDER
            if col_idx == 2:
                cell.fill = SERIAL_FILL
                cell.font = Font(bold=True)
                cell.alignment = CENTER_ALIGN
            elif col_idx in NUMERIC_COLUMNS:
                cell.alignment = CENTER_ALIGN
            else:
                cell.alignment = TEXT_ALIGN


def _render_gaps(ws, normalized: dict) -> None:
    gaps = normalized.get("coverage_gaps", [])
    for idx, gap in enumerate(gaps, start=1):
        excel_row = idx + 3
        ws.cell(row=excel_row, column=2, value=gap.get("leaf_id"))
        ws.cell(row=excel_row, column=3, value=gap.get("role"))
        ws.cell(row=excel_row, column=4, value=gap.get("axis_combo"))
        ws.cell(row=excel_row, column=5, value=gap.get("severity_estimate"))
    queue = normalized.get("confirmation_queue", [])
    for idx, row in enumerate(queue, start=1):
        excel_row = idx + 11
        ws.cell(row=excel_row, column=2, value=row.get("row_id"))
        ws.cell(row=excel_row, column=3, value=row.get("leaf_id"))
        ws.cell(row=excel_row, column=4, value=row.get("failure_mode"))
        ws.cell(row=excel_row, column=5, value=row.get("evidence_grade"))
        ws.cell(row=excel_row, column=6, value=row.get("confidence"))
        ws.cell(row=excel_row, column=7, value=row.get("rpn"))
        ws.cell(row=excel_row, column=8, value=round(row.get("confidence", 0) * row.get("rpn", 0), 2))


def _render_structure(ws, structure: dict | None, normalized: dict | None = None) -> None:
    def _walk(node, depth, lines):
        if not node:
            return
        lines.append(("  " * depth) + f"{node['id']} {node['name']} ({node['level']})")
        for c in node.get("children", []):
            _walk(c, depth + 1, lines)

    if structure:
        lines = []
        _walk(structure.get("hierarchy"), 0, lines)
        for i, line in enumerate(lines, start=1):
            ws.cell(row=2 + i, column=2, value=line)
        start = 2 + len(lines) + 3
        ws.cell(row=start, column=2, value="P-Diagrams")
        offset = 1
        for pd in structure.get("p_diagrams", []):
            ws.cell(row=start + offset, column=2, value=f"scope_id={pd['scope_id']}")
            offset += 1
            for axis in ("input_signals", "control_factors", "intended_outputs", "unintended_outputs", "error_states"):
                ws.cell(row=start + offset, column=2, value=f"  {axis}: {', '.join(pd.get(axis, []))}")
                offset += 1
            for sub_axis in pd.get("noise_factors", {}):
                ws.cell(row=start + offset, column=2, value=f"  noise.{sub_axis}: {', '.join(pd['noise_factors'][sub_axis])}")
                offset += 1
    else:
        # No structure provided: write scope paths from normalized rows as fallback hierarchy
        rows = (normalized or {}).get("rows", [])
        seen_paths = []
        for row in rows:
            path = row.get("scope_path", "")
            if path and path not in seen_paths:
                seen_paths.append(path)
        if seen_paths:
            ws.cell(row=3, column=2, value="(no structure.json provided — scope paths from normalized rows:)")
            for i, path in enumerate(seen_paths, start=1):
                ws.cell(row=3 + i, column=2, value=path)
        else:
            ws.cell(row=3, column=2, value="(no structure provided)")


def render(normalized_path: Path, structure_path: Path | None, output: Path) -> None:
    normalized = json.loads(normalized_path.read_text(encoding="utf-8"))
    structure = json.loads(structure_path.read_text(encoding="utf-8")) if structure_path else None

    shutil.copy(TEMPLATE, output)
    wb = load_workbook(output)
    _render_cover(wb["封面"], normalized, structure)
    _render_main(wb["FMEA主表"], normalized.get("rows", []))
    _render_gaps(wb["覆盖盲区与待确认队列"], normalized)
    _render_structure(wb["结构与P-Diagram"], structure, normalized)
    wb.save(output)


def main() -> None:
    parser = argparse.ArgumentParser(description="Render fmea_normalized.json to xlsx using template.xlsx.")
    parser.add_argument("--normalized", required=True, type=Path)
    parser.add_argument("--structure", type=Path, help="Optional structure.json for cover and Sheet 5.")
    parser.add_argument("--output", required=True, type=Path)
    args = parser.parse_args()
    render(args.normalized, args.structure, args.output)


if __name__ == "__main__":
    main()
