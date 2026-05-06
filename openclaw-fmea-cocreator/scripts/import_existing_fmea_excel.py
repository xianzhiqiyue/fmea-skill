from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


SCRIPT_DIR = Path(__file__).resolve().parent
SKILL_DIR = SCRIPT_DIR.parent
if str(SCRIPT_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPT_DIR))

import draft_fmea_from_cases as draft


SUMMARY_SHEETS = {"概览", "Scope规划", "确认队列", "Top风险", "建议动作", "来源追踪"}
STATUS_ALIASES = {
    "draft": "draft",
    "草稿": "draft",
    "needs expert confirmation": "needs expert confirmation",
    "待确认": "needs expert confirmation",
    "需要专家确认": "needs expert confirmation",
    "confirmed": "confirmed",
    "已确认": "confirmed",
}
REFERENCE_TYPE_ALIASES = {
    "current module": "current module",
    "当前模块": "current module",
    "direct family reference": "direct family reference",
    "直接家族参考": "direct family reference",
    "家族参考": "direct family reference",
    "broader analogy": "broader analogy",
    "更广泛类比": "broader analogy",
}
IMPORT_HEADER_ALIASES = {
    "scope": ["Scope", "分析范围", "scope", "子系统", "子系统/功能模块", "子系统/组件", "生命周期维度"],
    "analysis_object": ["Analysis object", "零件名称", "模块/零件", *draft.FIELD_ALIASES["analysis_object"]],
    "function": ["Function or requirement", "功能及要求", *draft.FIELD_ALIASES["function"]],
    "failure_mode": ["Failure mode", "潜在失效模式", *draft.FIELD_ALIASES["failure_mode"]],
    "effect": ["Failure effect", "失效影响（后果）", "失效影响(后果)", *draft.FIELD_ALIASES["effect"]],
    "severity": ["S", "Severity", "严重度 S", *draft.FIELD_ALIASES["severity"]],
    "cause": ["Cause or mechanism", "Cause", "失效原因", *draft.FIELD_ALIASES["cause"]],
    "occurrence": ["O", "Occurrence", "频度 O", *draft.FIELD_ALIASES["occurrence"]],
    "current_controls": ["Current controls", "Current control", "现行预防措施", "现行探测控制", *draft.FIELD_ALIASES["current_controls"]],
    "detection": ["D", "Detection", "探测度 D", *draft.FIELD_ALIASES["detection"]],
    "rpn": ["RPN", "Current RPN", *draft.FIELD_ALIASES["rpn"]],
    "recommended_actions": ["Recommended actions", "Recommended action", "建议措施", *draft.FIELD_ALIASES["recommended_actions"]],
    "post_action_severity": ["Post-action S", "改进后S", "措施后 S", *draft.FIELD_ALIASES["post_action_severity"]],
    "post_action_occurrence": ["Post-action O", "改进后O", "措施后 O", *draft.FIELD_ALIASES["post_action_occurrence"]],
    "post_action_detection": ["Post-action D", "改进后D", "措施后 D", *draft.FIELD_ALIASES["post_action_detection"]],
    "post_action_rpn": ["Post-action RPN", "改进后RPN", "措施后 RPN", *draft.FIELD_ALIASES["post_action_rpn"]],
    "owner": ["Owner", "措施负责人", *draft.FIELD_ALIASES["owner"]],
    "target_date": ["Target date", "Target Date", "完成时间", *draft.FIELD_ALIASES["target_date"]],
    "confirmation_status": ["Confirmation status", "确认状态", "评审状态", "状态"],
    "review_comment": ["Review comment", "评审备注", "评审说明", "备注", "确认备注"],
    "rating_basis": ["Rating basis", "评分依据", "打分依据", "AI打分推导依据"],
    "reference_type": ["Reference type", "来源类型", "参考类型", "引用类型"],
    "source_case": ["Source case", "Source cases", "来源案例", "来源", "案例来源", "追溯来源"],
}
SCOPE_PLAN_HEADERS = {
    "scope": ["Scope"],
    "query_terms": ["检索关键词", "关键词"],
    "source": ["来源"],
    "hit_count": ["命中数"],
    "reason": ["说明"],
}
CONFIRMATION_HEADERS = {
    "scope": ["Scope"],
    "row_key": ["Row key"],
    "why_confirmation_is_needed": ["Why confirmation is needed"],
    "suggested_reviewer_focus": ["Suggested reviewer focus"],
    "review_comment": ["Review comment"],
    "reference_type": ["Reference type"],
    "source_case": ["Source case"],
}


def normalize_header(value: Any) -> str:
    text = draft.normalize_space(str(value or ""))
    if not text:
        return ""
    return text.lower()


def normalize_status(value: str) -> str:
    cleaned = draft.normalize_space(value).lower()
    return STATUS_ALIASES.get(cleaned, value.strip())


def normalize_reference_type(value: str) -> str:
    cleaned = draft.normalize_space(value).lower()
    return REFERENCE_TYPE_ALIASES.get(cleaned, value.strip() or "current module")


def cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        return str(int(value)) if value.is_integer() else str(value)
    text = str(value).replace("\r\n", "\n").replace("\r", "\n")
    lines = [line.strip() for line in text.split("\n")]
    return "\n".join(line for line in lines if line)


def split_source_cases(value: str, import_trace: str) -> list[str]:
    raw_parts = re.split(r"[;\n]+", value)
    parts = [draft.normalize_space(part) for part in raw_parts if draft.normalize_space(part)]
    source_cases = [import_trace]
    for part in parts:
        if part not in source_cases:
            source_cases.append(part)
    return source_cases


def append_review_comment(existing: str, new_comment: str) -> str:
    current = [line.strip() for line in str(existing or "").splitlines() if line.strip()]
    cleaned = draft.normalize_space(new_comment)
    if cleaned and cleaned not in current:
        current.append(cleaned)
    return "\n".join(current)


def build_header_map(row_values: list[Any], aliases: dict[str, list[str]]) -> dict[str, int]:
    normalized_aliases = {
        field: {normalize_header(alias) for alias in alias_values}
        for field, alias_values in aliases.items()
    }
    mapping: dict[str, int] = {}
    for index, value in enumerate(row_values, start=1):
        header = normalize_header(value)
        if not header:
            continue
        for field, alias_set in normalized_aliases.items():
            if header in alias_set and field not in mapping:
                mapping[field] = index
    return mapping


def detect_header_row(ws: Any, aliases: dict[str, list[str]], max_scan_rows: int = 12) -> tuple[int, dict[str, int]] | None:
    best: tuple[int, int, dict[str, int]] | None = None
    for row_index, row in enumerate(ws.iter_rows(min_row=1, max_row=min(ws.max_row, max_scan_rows), values_only=True), start=1):
        mapping = build_header_map(list(row), aliases)
        score = len(mapping)
        if {"failure_mode", "analysis_object"} <= set(mapping):
            score += 2
        elif "failure_mode" in mapping and "function" in mapping:
            score += 1
        if best is None or score > best[0]:
            best = (score, row_index, mapping)
    if best is None or best[0] < 4:
        return None
    return best[1], best[2]


def parse_overview(workbook: Any) -> dict[str, str]:
    if "概览" not in workbook.sheetnames and "封面" not in workbook.sheetnames:
        return {}
    if "封面" in workbook.sheetnames:
        ws = workbook["封面"]
        overview: dict[str, str] = {}
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 24), min_col=2, max_col=3, values_only=True):
            key = draft.normalize_space(str(row[0] or ""))
            value = cell_text(row[1])
            if key and value:
                overview[key] = value
        return overview
    ws = workbook["概览"]
    overview: dict[str, str] = {}
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 20), min_col=1, max_col=2, values_only=True):
        key = draft.normalize_space(str(row[0] or ""))
        value = cell_text(row[1])
        if key and value:
            overview[key] = value
    return overview


def parse_scope_plan(workbook: Any) -> dict[str, draft.ScopeDefinition]:
    if "Scope规划" not in workbook.sheetnames:
        return {}
    ws = workbook["Scope规划"]
    detected = detect_header_row(ws, SCOPE_PLAN_HEADERS, max_scan_rows=5)
    if detected is None:
        return {}
    header_row, mapping = detected
    scope_plan: dict[str, draft.ScopeDefinition] = {}
    for row_index in range(header_row + 1, ws.max_row + 1):
        name = cell_text(ws.cell(row=row_index, column=mapping["scope"]).value) if mapping.get("scope") else ""
        if not name:
            continue
        query_value = cell_text(ws.cell(row=row_index, column=mapping["query_terms"]).value) if mapping.get("query_terms") else ""
        source_value = cell_text(ws.cell(row=row_index, column=mapping["source"]).value) if mapping.get("source") else ""
        hit_count_value = cell_text(ws.cell(row=row_index, column=mapping["hit_count"]).value) if mapping.get("hit_count") else ""
        reason_value = cell_text(ws.cell(row=row_index, column=mapping["reason"]).value) if mapping.get("reason") else ""
        query_terms = re.split(r"[/;\n, ]+", query_value)
        scope_plan[name] = draft.ScopeDefinition(
            name=name,
            query_terms=[term for term in (draft.normalize_space(term) for term in query_terms) if term],
            extracted_terms=[],
            auto_suggested=source_value.lower() == "auto",
            hit_count=draft.safe_int(hit_count_value) or 0,
            reason=reason_value,
        )
    return scope_plan


def split_reason_text(value: str) -> list[str]:
    parts = [draft.normalize_space(part) for part in re.split(r"[；;]+", value) if draft.normalize_space(part)]
    unique: list[str] = []
    for part in parts:
        if part not in unique:
            unique.append(part)
    return unique


def parse_boundary_scopes(reasons: list[str]) -> list[str]:
    boundary_scopes: list[str] = []
    for reason in reasons:
        if "也可能属于：" not in reason:
            continue
        _, raw_scopes = reason.split("也可能属于：", 1)
        for scope_name in re.split(r"[／/]+", raw_scopes):
            cleaned = draft.normalize_space(scope_name)
            if cleaned and cleaned not in boundary_scopes:
                boundary_scopes.append(cleaned)
    return boundary_scopes


def parse_confirmation_queue(workbook: Any) -> dict[tuple[str, str], dict[str, Any]]:
    if "确认队列" not in workbook.sheetnames:
        return {}
    ws = workbook["确认队列"]
    detected = detect_header_row(ws, CONFIRMATION_HEADERS, max_scan_rows=5)
    if detected is None:
        return {}
    header_row, mapping = detected
    queue_lookup: dict[tuple[str, str], dict[str, Any]] = {}
    for row_index in range(header_row + 1, ws.max_row + 1):
        scope = cell_text(ws.cell(row=row_index, column=mapping["scope"]).value) if mapping.get("scope") else ""
        row_key = cell_text(ws.cell(row=row_index, column=mapping["row_key"]).value) if mapping.get("row_key") else ""
        if not scope or not row_key:
            continue
        why_text = cell_text(ws.cell(row=row_index, column=mapping["why_confirmation_is_needed"]).value) if mapping.get("why_confirmation_is_needed") else ""
        queue_lookup[(scope, row_key)] = {
            "confirmation_reasons": split_reason_text(why_text),
            "reviewer_focus": cell_text(ws.cell(row=row_index, column=mapping["suggested_reviewer_focus"]).value) if mapping.get("suggested_reviewer_focus") else "",
            "review_comment": cell_text(ws.cell(row=row_index, column=mapping["review_comment"]).value) if mapping.get("review_comment") else "",
            "boundary_scopes": parse_boundary_scopes(split_reason_text(why_text)),
        }
    return queue_lookup


def sheet_scope_name(sheet_title: str) -> str:
    return re.sub(r"^\d+\s*[-_]\s*", "", draft.normalize_space(sheet_title))


def infer_confirmation_reasons(row: draft.DraftRow, explicit_status: str) -> list[str]:
    if explicit_status == "confirmed":
        return []

    reasons: list[str] = []
    missing_scores = [label for label, value in [("S", row.severity), ("O", row.occurrence), ("D", row.detection)] if not value]
    if missing_scores:
        reasons.append(f"源案例缺少完整评分字段：{'/'.join(missing_scores)}")

    if row.reference_type != "current module":
        reasons.append("该行主要来自跨模块家族类比，需要确认是否适用于当前模块责任范围")

    if explicit_status == "needs expert confirmation" and not reasons:
        reasons.append("该行来自已有 FMEA 导入，需补充明确的评审理由或确认记录")

    return reasons


def infer_rating_basis(row: draft.DraftRow, explicit_basis: str) -> str:
    if explicit_basis:
        return explicit_basis
    if row.severity and row.occurrence and row.detection:
        return "由已有 FMEA 导入，需结合原评审记录核验 S/O/D 和现行控制。"
    return "由已有 FMEA 导入，但评分字段或依据不完整，需补充确认。"


def build_import_context(args: argparse.Namespace, overview: dict[str, str]) -> str:
    if args.context_file:
        return Path(args.context_file).read_text(encoding="utf-8").strip()
    if args.context_text:
        return args.context_text.strip()

    summary_lines = [
        f"导入已有 FMEA 工作簿: {Path(args.input_excel).name}",
    ]
    if args.module:
        summary_lines.append(f"模块/分析对象名称:\n{args.module}")
    elif overview.get("模块"):
        summary_lines.append(f"模块/分析对象名称:\n{overview['模块']}")
    if args.fmea_type:
        summary_lines.append(f"FMEA 类型:\n{args.fmea_type}")
    elif overview.get("FMEA 类型"):
        summary_lines.append(f"FMEA 类型:\n{overview['FMEA 类型']}")
    return "\n\n".join(summary_lines)


def parse_scope_sheet(
    ws: Any,
    workbook_name: str,
    queue_lookup: dict[tuple[str, str], dict[str, Any]],
) -> tuple[str, list[draft.DraftRow]] | None:
    if ws.title in SUMMARY_SHEETS:
        return None
    detected = detect_header_row(ws, IMPORT_HEADER_ALIASES)
    if detected is None:
        return None

    header_row, mapping = detected
    default_scope = sheet_scope_name(ws.title)
    rows: list[draft.DraftRow] = []
    for row_index in range(header_row + 1, ws.max_row + 1):
        raw_values = {
            field: cell_text(ws.cell(row=row_index, column=column_index).value)
            for field, column_index in mapping.items()
        }
        if not any(raw_values.values()):
            continue

        if not raw_values.get("failure_mode") and raw_values.get("recommended_actions", "").startswith("未召回到足够案例"):
            continue
        if not raw_values.get("failure_mode") and not raw_values.get("analysis_object") and not raw_values.get("function"):
            continue

        scope = raw_values.get("scope") or default_scope
        import_trace = f"{workbook_name} / {ws.title} / row {row_index}"
        explicit_status = normalize_status(raw_values.get("confirmation_status", ""))
        reference_type = normalize_reference_type(raw_values.get("reference_type", ""))
        draft_row = draft.DraftRow(
            scope=scope,
            analysis_object=raw_values.get("analysis_object", ""),
            function=raw_values.get("function", ""),
            failure_mode=raw_values.get("failure_mode", ""),
            effect=raw_values.get("effect", ""),
            severity=raw_values.get("severity", ""),
            cause=raw_values.get("cause", ""),
            occurrence=raw_values.get("occurrence", ""),
            current_controls=raw_values.get("current_controls", ""),
            detection=raw_values.get("detection", ""),
            rpn=draft.compute_rpn(
                raw_values.get("severity", ""),
                raw_values.get("occurrence", ""),
                raw_values.get("detection", ""),
                raw_values.get("rpn", ""),
            ),
            recommended_actions=raw_values.get("recommended_actions", ""),
            owner=raw_values.get("owner", ""),
            target_date=raw_values.get("target_date", ""),
            confirmation_status="draft",
            rating_basis="",
            reference_type=reference_type,
            source_cases=split_source_cases(raw_values.get("source_case", ""), import_trace),
            review_comment=raw_values.get("review_comment", ""),
            post_action_severity=raw_values.get("post_action_severity", ""),
            post_action_occurrence=raw_values.get("post_action_occurrence", ""),
            post_action_detection=raw_values.get("post_action_detection", ""),
            post_action_rpn=draft.compute_post_action_rpn(
                raw_values.get("post_action_severity", ""),
                raw_values.get("post_action_occurrence", ""),
                raw_values.get("post_action_detection", ""),
                raw_values.get("post_action_rpn", ""),
            ),
            max_match_score=0,
            max_scope_hits=0,
            confirmation_reasons=[],
            reviewer_focus="",
            boundary_scopes=[],
        )

        row_lookup_key = (draft_row.scope, draft.row_key(draft_row))
        queue_item = queue_lookup.get(row_lookup_key, {})
        draft_row.confirmation_reasons = queue_item.get("confirmation_reasons") or infer_confirmation_reasons(draft_row, explicit_status)
        draft_row.boundary_scopes = queue_item.get("boundary_scopes", [])
        draft_row.rating_basis = infer_rating_basis(draft_row, raw_values.get("rating_basis", ""))
        draft_row.review_comment = append_review_comment(draft_row.review_comment, queue_item.get("review_comment", ""))
        if explicit_status == "confirmed":
            draft_row.confirmation_status = "confirmed"
        elif explicit_status == "needs expert confirmation" or queue_item:
            draft_row.confirmation_status = "needs expert confirmation"
        else:
            draft_row.confirmation_status = draft.build_confirmation_status(draft_row.confirmation_reasons)
        if draft_row.confirmation_status == "needs expert confirmation":
            draft_row.reviewer_focus = queue_item.get("reviewer_focus", "")
            if not draft_row.reviewer_focus:
                draft_row.reviewer_focus = draft.build_reviewer_focus(
                    draft_row.reference_type,
                    draft_row.boundary_scopes,
                    draft_row.occurrence,
                    draft_row.detection,
                )
            if not draft_row.reviewer_focus and draft_row.confirmation_reasons:
                draft_row.reviewer_focus = "补充确认评分依据、类比适用性和责任边界。"

        rows.append(draft_row)

    if not rows:
        return None
    rows.sort(
        key=lambda item: (
            -(draft.safe_int(item.rpn) or -1),
            item.analysis_object,
            item.failure_mode,
        )
    )
    return default_scope, rows


def build_scopes(scope_plan: dict[str, draft.ScopeDefinition], scope_rows: dict[str, list[draft.DraftRow]]) -> list[draft.ScopeDefinition]:
    scopes: list[draft.ScopeDefinition] = []
    for scope_name, scope in scope_plan.items():
        scopes.append(scope)
        scope_rows.setdefault(scope_name, [])
    for scope_name in scope_rows:
        if scope_name not in scope_plan:
            scopes.append(
                draft.ScopeDefinition(
                    name=scope_name,
                    query_terms=[],
                    extracted_terms=[],
                    auto_suggested=False,
                    hit_count=len(scope_rows[scope_name]),
                    reason="imported from existing workbook scope sheet",
                )
            )
    return scopes


def infer_module_and_type(args: argparse.Namespace, overview: dict[str, str]) -> tuple[str, str]:
    module = args.module or overview.get("模块", "") or overview.get("模块/分析对象名称", "") or overview.get("产品型号", "")
    fmea_type = args.fmea_type or overview.get("FMEA 类型", "") or "DFMEA"
    if not module:
        raise ValueError("无法从参数或工作簿概览中识别模块名，请提供 --module。")
    if fmea_type not in {"AFMEA", "SFMEA", "DFMEA"}:
        raise ValueError("无法从参数或工作簿概览中识别 FMEA 类型，请提供 --fmea-type。")
    return module, fmea_type


def main() -> None:
    parser = argparse.ArgumentParser(description="Import an existing FMEA Excel workbook into the OpenClaw review pipeline.")
    parser.add_argument("--input-excel", required=True, help="Path to the existing FMEA workbook.")
    parser.add_argument("--module", help="Optional module/analysis object name. Falls back to workbook overview if omitted.")
    parser.add_argument("--fmea-type", help="Optional FMEA type. Falls back to workbook overview if omitted.")
    parser.add_argument("--context-file", help="Optional UTF-8 text file describing this review/import context.")
    parser.add_argument("--context-text", help="Optional inline context for this imported review.")
    parser.add_argument("--excel-out", required=True, help="Path to the normalized Excel workbook.")
    parser.add_argument("--markdown-out", help="Optional path to save the Markdown preview.")
    parser.add_argument("--json-out", help="Optional path to save the JSON payload.")
    args = parser.parse_args()

    workbook_path = Path(args.input_excel)
    workbook = load_workbook(workbook_path, data_only=True)
    overview = parse_overview(workbook)
    module, fmea_type = infer_module_and_type(args, overview)
    context_text = build_import_context(args, overview)

    scope_plan = parse_scope_plan(workbook)
    queue_lookup = parse_confirmation_queue(workbook)
    scope_rows: dict[str, list[draft.DraftRow]] = {}
    for ws in workbook.worksheets:
        parsed = parse_scope_sheet(ws, workbook_path.name, queue_lookup)
        if parsed is None:
            continue
        _, rows = parsed
        for row in rows:
            scope_rows.setdefault(row.scope, []).append(row)

    if not scope_rows:
        raise ValueError("未在工作簿中识别到可导入的 FMEA 主表，请检查表头或工作表结构。")

    scopes = build_scopes(scope_plan, scope_rows)
    draft.render_excel_workbook(module, fmea_type, context_text, scopes, scope_rows, Path(args.excel_out))

    payload = draft.build_json_payload(module, fmea_type, context_text, scopes, scope_rows)
    payload["import_source"] = {
        "workbook_path": str(workbook_path),
        "workbook_name": workbook_path.name,
        "sheet_count": len(workbook.worksheets),
    }

    if args.markdown_out:
        markdown_path = Path(args.markdown_out)
        markdown_path.parent.mkdir(parents=True, exist_ok=True)
        markdown_path.write_text(draft.render_markdown(module, fmea_type, context_text, scopes, scope_rows), encoding="utf-8")

    if args.json_out:
        json_path = Path(args.json_out)
        json_path.parent.mkdir(parents=True, exist_ok=True)
        json_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


if __name__ == "__main__":
    main()
