from __future__ import annotations

import csv
import json
import re
import zipfile
from collections import defaultdict
from copy import deepcopy
from dataclasses import dataclass
from datetime import date, datetime, time
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent.parent
OUTPUT_ROOT = ROOT / "excel_materials"


THEME_META = {
    "dfmea_project_plan": {
        "name": "项目计划与推进",
        "description": "模块负责人、时间计划、试点推进和进度跟踪资料。",
    },
    "dfmea_sample_data": {
        "name": "DFMEA 成品样例",
        "description": "已经形成的 DFMEA 表格，可直接作为案例库、字段模板和输出样板。",
    },
    "ai_quality_strategy": {
        "name": "质量管理与 AI 赋能策略",
        "description": "质量管理环节、AI 赋能点、整体目标与实施背景。",
    },
    "ai_fmea_methodology": {
        "name": "AI-FMEA 方法与平台设计",
        "description": "流程设计、平台蓝图、输入要求、功能规划与实施方式。",
    },
    "ai_prompt_templates": {
        "name": "提示语与分析模板",
        "description": "AFMEA/SFMEA/DFMEA 提示语模板、失效分类和提问框架。",
    },
    "knowledge_base_template": {
        "name": "案例库与知识沉淀模板",
        "description": "适合转成知识库、案例库和标准化输入输出结构的模板资料。",
    },
}


WORKBOOK_CONFIG = {
    "CAN400产品DFMEA.xlsx": {
        "display_name": "CAN400产品DFMEA",
        "workbook_type": "DFMEA 样例库",
        "description": "NMR/CAN400 相关模块的 DFMEA 成品表和项目推进信息。",
        "sheet_meta": {
            "项目计划": {
                "theme": "dfmea_project_plan",
                "description": "模块负责人、进度，以及一段自然语言 DFMEA 生成示例。",
                "development_use": "可用于梳理模块清单、负责人映射，以及抽取自然语言输入样板。",
            },
            "变温系统": {
                "theme": "dfmea_sample_data",
                "description": "低温/制冷系统的完整 DFMEA 样例，含责任人和计划完成时间。",
                "development_use": "最适合作为第一批案例库和输出表头标准来源。",
            },
            "调谐单元": {
                "theme": "dfmea_sample_data",
                "description": "射频调谐、机械限位、EMC 与算法风险样例。",
                "development_use": "适合作为射频系统和机电控制类 DFMEA 参考。",
            },
            "自动进样器": {
                "theme": "dfmea_sample_data",
                "description": "自动进样器的运动、抓取、检测和气路 DFMEA。",
                "development_use": "适合作为机电一体化和样品搬运类场景案例。",
            },
            "收发机": {
                "theme": "dfmea_sample_data",
                "description": "收发机的发射、接收、混频、ADC 与连接风险样例。",
                "development_use": "适合作为电子与射频链路 DFMEA 参考。",
            },
            "前置放大器": {
                "theme": "dfmea_sample_data",
                "description": "前置放大器的保护、LNA、接口和模式逻辑风险样例。",
                "development_use": "适合作为前端保护和低噪声链路案例。",
            },
            "射频功放": {
                "theme": "dfmea_sample_data",
                "description": "射频功放的保护、放大、切换、热管理和复位逻辑样例。",
                "development_use": "包含全表最高风险项，适合做高优先级风险规则样板。",
            },
            "进样筒": {
                "theme": "dfmea_sample_data",
                "description": "进样筒结构、气动和转子旋转相关 DFMEA。",
                "development_use": "包含较多续行写法，适合验证表格结构化处理策略。",
            },
            "匀场单元": {
                "theme": "dfmea_sample_data",
                "description": "匀场线圈、恒流驱动、连接和算法 DFMEA。",
                "development_use": "适合作为高精度控制、EMC 和长期稳定性案例。",
            },
            "电子学机柜": {
                "theme": "dfmea_sample_data",
                "description": "机柜电气、结构、门板安全和热设计 DFMEA。",
                "development_use": "适合作为系统级集成、结构与散热风险案例。",
            },
        },
    },
    "AI质量赋能.xlsx": {
        "display_name": "AI质量赋能",
        "workbook_type": "AI-FMEA 方法与规划库",
        "description": "AI 赋能质量工作的总体思路、AI-FMEA 平台设计、计划推进与案例模板。",
        "sheet_meta": {
            "Sheet1": {
                "theme": "ai_quality_strategy",
                "description": "质量主要工作模块与 AI 赋能方法清单。",
                "development_use": "适合提炼 AI 能力边界和质量业务场景分类。",
            },
            "质量管理与策划大纲V1.0": {
                "theme": "ai_quality_strategy",
                "description": "质量管理环节、策划控制内容和 AI 赋能方式。",
                "development_use": "适合作为总体流程背景和 AI 覆盖范围定义。",
            },
            "Sheet2": {
                "theme": "ai_fmea_methodology",
                "description": "AI 为 FMEA 赋能的任务清单和难点说明。",
                "development_use": "适合作为第一版产品待办和人工校准边界说明。",
            },
            "AI-FMEA平台构建": {
                "theme": "ai_fmea_methodology",
                "description": "AI-FMEA 平台建设步骤、输出、负责人和时间。",
                "development_use": "适合转成项目里程碑和职责分工。",
            },
            "模型规划蓝图": {
                "theme": "ai_fmea_methodology",
                "description": "AI-FMEA Agent 的能力规划蓝图。",
                "development_use": "可直接转成 skill/workflow 的能力分层。",
            },
            "AI-FMEA提示语模板": {
                "theme": "ai_prompt_templates",
                "description": "AFMEA/SFMEA/DFMEA 与失效分类提示语模板。",
                "development_use": "适合作为 skill 的 prompt references。",
            },
            "NMR-FMEA计划与实施": {
                "theme": "dfmea_project_plan",
                "description": "NMR 模块的 FMEA 计划、成员、时间和备注。",
                "development_use": "适合建立模块清单与试点推进跟踪。",
            },
            "DB600模组-FMEA计划与实施": {
                "theme": "dfmea_project_plan",
                "description": "DB600 模组 DFMEA 计划以及建议提供资料清单。",
                "development_use": "适合作为输入 checklist 模板。",
            },
            "XRD-FMEA计划与实施": {
                "theme": "dfmea_project_plan",
                "description": "XRD FMEA 项目的简版计划信息。",
                "development_use": "适合作为跨产品线复用的排期样板。",
            },
            "量产品SW+FMEA+PSP计划和跟进表": {
                "theme": "dfmea_project_plan",
                "description": "量产品改善周、工具组合和对接安排。",
                "development_use": "说明 FMEA 可延伸到量产改善工作流。",
            },
            "案例库模板": {
                "theme": "knowledge_base_template",
                "description": "失效案例标准化模板与示例。",
                "development_use": "可直接作为案例库导入格式。",
            },
        },
    },
}


@dataclass
class SheetExport:
    sheet_name: str
    safe_name: str
    theme: str
    description: str
    development_use: str
    row_count: int
    column_count: int
    nonempty_row_count: int
    continuation_rows: list[int]
    headers: list[str]
    csv_raw: str
    csv_filled: str
    json_path: str
    markdown_path: str


def safe_name(name: str) -> str:
    return re.sub(r'[\\/:*?"<>|]+', "_", name).strip().replace(" ", "_")


def stringify(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.isoformat()
    text = str(value)
    return text.strip()


def trim_rows(rows: list[list[str]]) -> list[list[str]]:
    last_row = 0
    last_col = 0
    for row_idx, row in enumerate(rows, start=1):
        row_has_value = False
        for col_idx, cell in enumerate(row, start=1):
            if cell != "":
                row_has_value = True
                last_col = max(last_col, col_idx)
        if row_has_value:
            last_row = row_idx
    trimmed = [row[:last_col] for row in rows[:last_row]]
    return trimmed


def fill_down_leading_blanks(rows: list[list[str]]) -> list[list[str]]:
    filled = deepcopy(rows)
    if not filled:
        return filled
    previous = filled[0]
    for row in filled[1:]:
        try:
            first_nonempty = next(i for i, cell in enumerate(row) if cell != "")
        except StopIteration:
            previous = row
            continue
        for index in range(first_nonempty):
            if row[index] == "" and previous[index] != "":
                row[index] = previous[index]
        previous = row
    return filled


def continuation_row_numbers(rows: list[list[str]]) -> list[int]:
    numbers: list[int] = []
    for index, row in enumerate(rows[1:], start=2):
        if not any(cell != "" for cell in row):
            continue
        try:
            first_nonempty = next(i for i, cell in enumerate(row) if cell != "")
        except StopIteration:
            continue
        if first_nonempty > 0:
            numbers.append(index)
    return numbers


def make_row_objects(headers: list[str], rows: list[list[str]]) -> list[dict[str, str]]:
    normalized_headers: list[str] = []
    used: dict[str, int] = defaultdict(int)
    for col_index, header in enumerate(headers, start=1):
        key = header or f"__col_{col_index}"
        used[key] += 1
        if used[key] > 1:
            key = f"{key}_{used[key]}"
        normalized_headers.append(key)

    objects: list[dict[str, str]] = []
    for excel_row_number, row in enumerate(rows[1:], start=2):
        if not any(cell != "" for cell in row):
            continue
        payload = {"__excel_row__": str(excel_row_number)}
        for key, cell in zip(normalized_headers, row):
            payload[key] = cell
        objects.append(payload)
    return objects


def csv_write(path: Path, rows: list[list[str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as file:
        writer = csv.writer(file)
        writer.writerows(rows)


def json_write(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as file:
        json.dump(payload, file, ensure_ascii=False, indent=2)


def md_escape(text: str) -> str:
    return text.replace("|", "\\|").replace("\n", "<br>")


def markdown_table(headers: list[str], rows: list[list[str]]) -> str:
    normalized_headers = ["Excel行号", *[header or "" for header in headers]]
    lines = [
        "| " + " | ".join(md_escape(item) for item in normalized_headers) + " |",
        "| " + " | ".join(["---"] * len(normalized_headers)) + " |",
    ]
    for excel_row_number, row in enumerate(rows[1:], start=2):
        if not any(cell != "" for cell in row):
            continue
        rendered = [str(excel_row_number), *row]
        lines.append("| " + " | ".join(md_escape(item) for item in rendered) + " |")
    return "\n".join(lines)


def extract_media(workbook_path: Path, target_dir: Path) -> list[str]:
    target_dir.mkdir(parents=True, exist_ok=True)
    exported: list[str] = []
    with zipfile.ZipFile(workbook_path) as archive:
        for member in archive.namelist():
            if not member.startswith("xl/media/"):
                continue
            filename = Path(member).name
            output_path = target_dir / filename
            output_path.write_bytes(archive.read(member))
            exported.append(str(output_path.relative_to(OUTPUT_ROOT)))
    return exported


def build_sheet_markdown(
    workbook_name: str,
    workbook_path: Path,
    export: SheetExport,
    headers: list[str],
    raw_rows: list[list[str]],
    filled_rows: list[list[str]],
) -> str:
    theme_meta = THEME_META[export.theme]
    header_lines = [
        f"# {export.sheet_name}",
        "",
        f"- 所属工作簿: `{workbook_name}`",
        f"- 原始文件: `{workbook_path}`",
        f"- 分类: `{theme_meta['name']}`",
        f"- 工作表说明: {export.description}",
        f"- 对后续开发的作用: {export.development_use}",
        f"- 表头列数: {export.column_count}",
        f"- 有效行数（含表头）: {export.row_count}",
        f"- 非空数据行数（不含表头）: {export.nonempty_row_count}",
        f"- 续行/继承上下文行号: {', '.join(str(num) for num in export.continuation_rows) if export.continuation_rows else '无'}",
        "",
        "## 导出文件",
        "",
        f"- 原始 CSV: `{export.csv_raw}`",
        f"- 续行填充版 CSV: `{export.csv_filled}`",
        f"- 结构化 JSON: `{export.json_path}`",
        "",
        "## 表头",
        "",
    ]

    header_table = [
        "| 列序号 | 列名 |",
        "| --- | --- |",
    ]
    for index, header in enumerate(headers, start=1):
        header_table.append(f"| {index} | {md_escape(header)} |")

    body = [
        "",
        "## 续行填充后的完整表格",
        "",
        "> 说明：为便于阅读，下面的表格对前导空白单元格做了“沿用上一行上下文”的填充。原始空白保留在 CSV/JSON 中。",
        "",
        markdown_table(headers, filled_rows),
    ]
    return "\n".join(header_lines + header_table + body) + "\n"


def build_workbook_overview(
    workbook_name: str,
    workbook_path: Path,
    workbook_meta: dict[str, Any],
    exports: list[SheetExport],
    media_paths: list[str],
) -> str:
    lines = [
        f"# {workbook_meta['display_name']}",
        "",
        f"- 原始文件: `{workbook_path}`",
        f"- 工作簿类型: {workbook_meta['workbook_type']}",
        f"- 定位说明: {workbook_meta['description']}",
        f"- 工作表数量: {len(exports)}",
        "",
        "## 工作表索引",
        "",
        "| 工作表 | 分类 | 非空数据行 | 导出 Markdown | 说明 |",
        "| --- | --- | ---: | --- | --- |",
    ]
    for export in exports:
        lines.append(
            "| "
            + " | ".join(
                [
                    md_escape(export.sheet_name),
                    md_escape(THEME_META[export.theme]["name"]),
                    str(export.nonempty_row_count),
                    md_escape(export.markdown_path),
                    md_escape(export.description),
                ]
            )
            + " |"
        )

    if media_paths:
        lines.extend(
            [
                "",
                "## 嵌入媒体",
                "",
                "| 文件 |",
                "| --- |",
            ]
        )
        for media_path in media_paths:
            lines.append(f"| `{media_path}` |")
    return "\n".join(lines) + "\n"


def build_root_readme(
    workbook_index: list[dict[str, Any]],
    theme_index: dict[str, list[dict[str, str]]],
) -> str:
    lines = [
        "# Excel 资料整理索引",
        "",
        "本目录是基于原始 Excel 自动生成的资料整理结果，目标是为后续的 OpenClaw/FMEA 开发工作提供稳定输入。",
        "",
        "## 原始文件保留说明",
        "",
        "- 根目录下的原始 Excel 文件保留不动，没有删除、改名或覆盖。",
        f"- 原始文件 1: `{ROOT / 'CAN400产品DFMEA.xlsx'}`",
        f"- 原始文件 2: `{ROOT / 'AI质量赋能.xlsx'}`",
        "",
        "## 目录结构",
        "",
        "- `workbooks/`: 按工作簿拆分的逐表导出结果。",
        "- `theme_index.md`: 按主题分类的索引。",
        "",
        "## 工作簿索引",
        "",
        "| 工作簿 | 类型 | 说明 | 总览文件 |",
        "| --- | --- | --- | --- |",
    ]
    for item in workbook_index:
        lines.append(
            "| "
            + " | ".join(
                [
                    md_escape(item["display_name"]),
                    md_escape(item["workbook_type"]),
                    md_escape(item["description"]),
                    md_escape(item["overview_path"]),
                ]
            )
            + " |"
        )

    lines.extend(
        [
            "",
            "## 主题分类索引",
            "",
            "| 分类 | 说明 | 工作表数量 |",
            "| --- | --- | ---: |",
        ]
    )
    for theme_key, items in theme_index.items():
        meta = THEME_META[theme_key]
        lines.append(
            f"| {meta['name']} | {meta['description']} | {len(items)} |"
        )

    lines.extend(
        [
            "",
            "## 建议使用方式",
            "",
            "1. 先看 `workbooks/*/workbook_overview.md` 确认每个工作簿和工作表的作用。",
            "2. 需要完整信息时，优先查看对应 sheet 的 Markdown 和 CSV。",
            "3. 做程序开发、索引构建或知识库导入时，优先消费 JSON 与 CSV。",
        ]
    )
    return "\n".join(lines) + "\n"


def build_theme_index(theme_index: dict[str, list[dict[str, str]]]) -> str:
    lines = [
        "# 主题分类索引",
        "",
        "下面按资料用途对两个 Excel 的工作表进行分类，便于后续开发按主题取用。",
    ]
    for theme_key, items in theme_index.items():
        meta = THEME_META[theme_key]
        lines.extend(
            [
                "",
                f"## {meta['name']}",
                "",
                meta["description"],
                "",
                "| 工作簿 | 工作表 | 导出 Markdown | 说明 | 开发用途 |",
                "| --- | --- | --- | --- | --- |",
            ]
        )
        for item in items:
            lines.append(
                "| "
                + " | ".join(
                    [
                        md_escape(item["workbook"]),
                        md_escape(item["sheet"]),
                        md_escape(item["markdown_path"]),
                        md_escape(item["description"]),
                        md_escape(item["development_use"]),
                    ]
                )
                + " |"
            )
    return "\n".join(lines) + "\n"


def main() -> None:
    OUTPUT_ROOT.mkdir(parents=True, exist_ok=True)

    workbook_index: list[dict[str, Any]] = []
    theme_index: dict[str, list[dict[str, str]]] = defaultdict(list)

    for workbook_filename, workbook_meta in WORKBOOK_CONFIG.items():
        workbook_path = ROOT / workbook_filename
        workbook_dir = OUTPUT_ROOT / "workbooks" / safe_name(workbook_meta["display_name"])
        csv_dir = workbook_dir / "csv"
        filled_dir = workbook_dir / "csv_filled"
        json_dir = workbook_dir / "json"
        markdown_dir = workbook_dir / "sheets"
        media_dir = workbook_dir / "media"

        workbook = load_workbook(workbook_path, data_only=True)
        exports: list[SheetExport] = []

        for index, worksheet in enumerate(workbook.worksheets, start=1):
            raw_rows = [
                [stringify(cell) for cell in row]
                for row in worksheet.iter_rows(values_only=True)
            ]
            raw_rows = trim_rows(raw_rows)
            if not raw_rows:
                raw_rows = [[]]
            headers = raw_rows[0]
            filled_rows = fill_down_leading_blanks(raw_rows)
            safe_sheet_name = f"{index:02d}_{safe_name(worksheet.title)}"
            sheet_meta = workbook_meta["sheet_meta"].get(
                worksheet.title,
                {
                    "theme": "knowledge_base_template",
                    "description": "未显式标注的工作表。",
                    "development_use": "建议后续人工补充分组和用途说明。",
                },
            )

            raw_csv_path = csv_dir / f"{safe_sheet_name}.csv"
            filled_csv_path = filled_dir / f"{safe_sheet_name}.csv"
            json_path = json_dir / f"{safe_sheet_name}.json"
            markdown_path = markdown_dir / f"{safe_sheet_name}.md"

            csv_write(raw_csv_path, raw_rows)
            csv_write(filled_csv_path, filled_rows)
            json_write(
                json_path,
                {
                    "workbook": workbook_filename,
                    "sheet_name": worksheet.title,
                    "theme": sheet_meta["theme"],
                    "description": sheet_meta["description"],
                    "development_use": sheet_meta["development_use"],
                    "headers": headers,
                    "raw_rows": make_row_objects(headers, raw_rows),
                    "filled_rows": make_row_objects(headers, filled_rows),
                    "continuation_rows": continuation_row_numbers(raw_rows),
                },
            )

            export = SheetExport(
                sheet_name=worksheet.title,
                safe_name=safe_sheet_name,
                theme=sheet_meta["theme"],
                description=sheet_meta["description"],
                development_use=sheet_meta["development_use"],
                row_count=len(raw_rows),
                column_count=len(headers),
                nonempty_row_count=sum(1 for row in raw_rows[1:] if any(cell != "" for cell in row)),
                continuation_rows=continuation_row_numbers(raw_rows),
                csv_raw=str(raw_csv_path.relative_to(OUTPUT_ROOT)),
                csv_filled=str(filled_csv_path.relative_to(OUTPUT_ROOT)),
                json_path=str(json_path.relative_to(OUTPUT_ROOT)),
                markdown_path=str(markdown_path.relative_to(OUTPUT_ROOT)),
                headers=headers,
            )

            markdown_path.parent.mkdir(parents=True, exist_ok=True)
            markdown_path.write_text(
                build_sheet_markdown(
                    workbook_filename,
                    workbook_path,
                    export,
                    headers,
                    raw_rows,
                    filled_rows,
                ),
                encoding="utf-8",
            )
            exports.append(export)

            theme_index[export.theme].append(
                {
                    "workbook": workbook_meta["display_name"],
                    "sheet": export.sheet_name,
                    "markdown_path": export.markdown_path,
                    "description": export.description,
                    "development_use": export.development_use,
                }
            )

        media_paths = extract_media(workbook_path, media_dir)
        manifest_path = workbook_dir / "manifest.json"
        overview_path = workbook_dir / "workbook_overview.md"

        json_write(
            manifest_path,
            {
                "workbook": workbook_filename,
                "display_name": workbook_meta["display_name"],
                "workbook_type": workbook_meta["workbook_type"],
                "description": workbook_meta["description"],
                "overview_path": str(overview_path.relative_to(OUTPUT_ROOT)),
                "media": media_paths,
                "sheets": [export.__dict__ for export in exports],
            },
        )
        overview_path.write_text(
            build_workbook_overview(
                workbook_filename,
                workbook_path,
                workbook_meta,
                exports,
                media_paths,
            ),
            encoding="utf-8",
        )

        workbook_index.append(
            {
                "display_name": workbook_meta["display_name"],
                "workbook_type": workbook_meta["workbook_type"],
                "description": workbook_meta["description"],
                "overview_path": str(overview_path.relative_to(OUTPUT_ROOT)),
            }
        )

    (OUTPUT_ROOT / "README.md").write_text(
        build_root_readme(workbook_index, theme_index),
        encoding="utf-8",
    )
    (OUTPUT_ROOT / "theme_index.md").write_text(
        build_theme_index(theme_index),
        encoding="utf-8",
    )


if __name__ == "__main__":
    main()
