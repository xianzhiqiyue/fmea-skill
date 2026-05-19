"""Verify import_existing_fmea_excel.py emits 31-column workbook for legacy 22-col input."""
import json
import sys
import subprocess
from pathlib import Path
import pytest
from openpyxl import Workbook, load_workbook

REPO_ROOT = Path(__file__).resolve().parent.parent
SCRIPT = REPO_ROOT / "openclaw-fmea-cocreator" / "scripts" / "import_existing_fmea_excel.py"


def _make_legacy_fmea_fixture(path: Path) -> None:
    """Create a minimal 22-column legacy FMEA Excel workbook with data rows."""
    wb = Workbook()
    # 封面 sheet with module/type metadata
    cover = wb.active
    cover.title = "封面"
    cover["B2"] = "测试散热模块 DFMEA分析报告"
    cover["B6"] = "模块"
    cover["C6"] = "测试散热模块"
    cover["B7"] = "FMEA 类型"
    cover["C7"] = "DFMEA"

    # FMEA scope sheet with 22-column legacy layout and actual data rows
    ws = wb.create_sheet("散热系统")
    # Row 1: column headers matching IMPORT_HEADER_ALIASES
    ws.append([
        None, "序号", "生命周期维度", "模块/零件", "功能及要求", "参数指标性能",
        "失效影响（后果）", "严重度\nS", "潜在失效模式", "失效原因",
        "现行预防措施", "频度\nO", "现行探测控制", "探测度\nD", "RPN",
        "AI打分推导依据", "建议措施", "措施负责人", "完成时间",
        "改进后S", "改进后O", "改进后D", "改进后RPN",
    ])
    # Row 2: data row 1
    ws.append([
        None, 1, "散热系统", "散热器", "散热", "温度",
        "电子元件损坏", "7", "散热片堵塞", "灰尘积累",
        "定期清洗", "5", "温度传感器", "4", "140",
        "经验判断", "清洁散热通道", "工程师", "2025-12",
        "", "", "", "",
    ])
    # Row 3: data row 2
    ws.append([
        None, 2, "散热系统", "风扇", "强制对流", "风速",
        "过热关机", "8", "风扇停转", "电机故障",
        "振动监测", "4", "温度报警", "3", "96",
        "行业标准", "更换风扇电机", "维修工", "2025-11",
        "", "", "", "",
    ])
    wb.save(path)


def test_legacy_import_produces_31_columns(tmp_path):
    # Always synthesise a fixture with real data rows for deterministic testing.
    # (template_legacy.xlsx in the repo is an empty formatting template with no data rows.)
    legacy = tmp_path / "legacy_input.xlsx"
    _make_legacy_fmea_fixture(legacy)

    out_xlsx = tmp_path / "imported.xlsx"
    out_json = tmp_path / "imported.json"
    result = subprocess.run(
        [sys.executable, str(SCRIPT),
         "--input-excel", str(legacy),
         "--module", "测试散热模块",
         "--fmea-type", "DFMEA",
         "--excel-out", str(out_xlsx),
         "--json-out", str(out_json)],
        capture_output=True, text=True, cwd=REPO_ROOT,
    )
    assert result.returncode == 0, f"Script failed:\n{result.stderr}\n{result.stdout}"
    wb = load_workbook(out_xlsx)
    assert "FMEA主表" in wb.sheetnames, "Output workbook missing FMEA主表 sheet"
    ws = wb["FMEA主表"]
    headers = [ws.cell(row=2, column=c).value for c in range(2, 33)]
    assert len(headers) == 31, f"Expected 31 headers, got {len(headers)}: {headers}"
    assert "Evidence grade" in headers, f"'Evidence grade' not found in headers: {headers}"

    # Verify M2 defaults written to data rows
    if ws.max_row >= 3:
        # col 25 = Evidence grade should be "ai-inferred" for imported rows
        evidence_grade = ws.cell(row=3, column=25).value
        assert evidence_grade == "ai-inferred", (
            f"Expected 'ai-inferred' in col 25, got {evidence_grade!r}"
        )
        # col 30 = Needs human confirmation should be "Y"
        needs_confirm = ws.cell(row=3, column=30).value
        assert needs_confirm == "Y", (
            f"Expected 'Y' in col 30 (Needs human confirmation), got {needs_confirm!r}"
        )

