"""Test build_workbook.py renders fmea_normalized.json into the new 31-column template."""
import sys
import subprocess
import json
from pathlib import Path

import pytest
from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parent.parent
SCRIPT = REPO_ROOT / "openclaw-fmea-cocreator" / "scripts" / "build_workbook.py"
FIXTURES = REPO_ROOT / "tests" / "fixtures"


@pytest.fixture
def rendered_workbook(tmp_path):
    out = tmp_path / "out.xlsx"
    result = subprocess.run(
        [sys.executable, str(SCRIPT),
         "--normalized", str(FIXTURES / "sample_normalized.json"),
         "--output", str(out)],
        capture_output=True, text=True, cwd=REPO_ROOT,
    )
    assert result.returncode == 0, result.stderr
    return load_workbook(out)


def test_workbook_has_5_sheets(rendered_workbook):
    expected = {"封面", "FMEA主表", "评分准则参考", "覆盖盲区与待确认队列", "结构与P-Diagram"}
    assert set(rendered_workbook.sheetnames) == expected


def test_main_sheet_has_31_columns_and_correct_headers(rendered_workbook):
    ws = rendered_workbook["FMEA主表"]
    headers = [ws.cell(row=2, column=c).value for c in range(2, 33)]
    assert len(headers) == 31
    assert headers[5] == "P-Diagram 锚点"
    assert headers[7] == "Failure mode canonical"
    assert headers[23] == "Evidence grade"
    assert headers[24] == "Confidence"


def test_main_sheet_first_row_data(rendered_workbook):
    ws = rendered_workbook["FMEA主表"]
    assert ws.cell(row=3, column=2).value == 1
    assert ws.cell(row=3, column=8).value == "触点粘连"
    assert ws.cell(row=3, column=9).value == "stuck_relay_contact"
    assert ws.cell(row=3, column=11).value == 9
    assert ws.cell(row=3, column=17).value == 315 or str(ws.cell(row=3, column=17).value).startswith("=")
    assert ws.cell(row=3, column=25).value == "evidence-backed"


def test_rpn_formula_present(rendered_workbook):
    ws = rendered_workbook["FMEA主表"]
    cell = ws.cell(row=3, column=17)
    assert cell.value == 315 or str(cell.value).startswith("=")


def test_coverage_gaps_sheet_filled(rendered_workbook):
    ws = rendered_workbook["覆盖盲区与待确认队列"]
    assert ws.cell(row=4, column=2).value == "T.1.2"


def test_structure_sheet_has_hierarchy_text(rendered_workbook):
    ws = rendered_workbook["结构与P-Diagram"]
    found = False
    for row in ws.iter_rows(min_row=3, max_row=20, values_only=True):
        for cell in row:
            if cell and "T" in str(cell):
                found = True
                break
    assert found


def test_cover_sheet_filled(rendered_workbook):
    ws = rendered_workbook["封面"]
    title = ws["B2"].value or ""
    assert "测试模块" in title
    assert "DFMEA" in title
